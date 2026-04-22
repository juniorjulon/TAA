"""
seed_taa_config.py
==================
One-time seed script. Creates `config/taa_config.xlsx` from the current
authoritative sources (CLAUDE.md + docs/TAA_Signal_Methodology.html + src/config.py).

After this file exists, you (the user) own it. Edit the Excel directly to
add/modify signals, asset classes, and weights, then run:

    python src/build_dashboard.py

to push changes into index.html and src/config.py.

Sheets created
--------------
  0. Instructions          how to use, edit, add rows
  1. AssetClasses          12 ACs with id, labels, group, benchmark, color, max_tilt
  2. DataSeries            master catalogue of every signal (ticker, source, freq, pillar)
  3. PillarWeights         AC x {F, M, S, V} weights (sum = 1.0)
  4. SignalMapping         (ac_id, series_id, pillar) -> sign, weight_in_pillar, note
  5. PillarNotes           optional header note per (ac_id, pillar)
"""

from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
OUT  = os.path.join(ROOT, "config", "taa_config.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)

# Style tokens ----------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1E2E47")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
INSTR_FILL  = PatternFill("solid", fgColor="FDE68A")
WARN_FILL   = PatternFill("solid", fgColor="FEE2E2")
THIN        = Side(border_style="thin", color="D1D5DB")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _apply_header(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER

def _autosize(ws, min_w=10, max_w=55):
    for col in ws.columns:
        letter = col[0].column_letter
        longest = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[letter].width = max(min_w, min(max_w, longest + 2))

# =============================================================================
# SHEET 0 - INSTRUCTIONS
# =============================================================================
def _sheet_instructions(wb):
    ws = wb.create_sheet("Instructions")
    ws.sheet_properties.tabColor = "C41230"
    rows = [
        ["TAA Dashboard Configuration - User Guide", ""],
        ["", ""],
        ["This workbook is the single source of truth for the TAA dashboard.",""],
        ["Editing it and running `python src/build_dashboard.py` will regenerate",""],
        ["the methodology blueprint in index.html and update src/config.py.",""],
        ["", ""],
        ["SHEETS OVERVIEW", ""],
        ["AssetClasses",   "12 asset classes. Add new rows to extend the universe."],
        ["DataSeries",     "Catalogue of every data series (signal) you track."],
        ["PillarWeights",  "How much each pillar (F,M,S,V) contributes to the composite per AC."],
        ["SignalMapping",  "The core table: which signals apply to which AC, per pillar, with weights."],
        ["PillarNotes",    "Optional italic note rendered at the top of each pillar card."],
        ["", ""],
        ["EDIT RULES", ""],
        ["Do not rename column headers.",                  "The build script matches by header name."],
        ["AC IDs are snake_case and must be unique.",      "e.g. money_market, us_equity, china_equity."],
        ["Series IDs are snake_case and must be unique.",  "e.g. ism_mfg_pmi, vix_level, bbb_oas."],
        ["pillar must be one of F, M, S, V.",              "Uppercase single letter."],
        ["sign column uses: +, ++, -, --, C, n/a.",        "C = contrarian (extremes only)."],
        ["weight_in_pillar is a percent string ('35%').",  "Weights within each (ac, pillar) should sum to ~100%."],
        ["PillarWeights rows must sum to 1.00.",           "F + M + S + V = 1.00 per AC."],
        ["", ""],
        ["HOW TO ADD A NEW DATA SERIES", ""],
        ["1. Add row in `DataSeries` with series_id, name, ticker, source, frequency, pillar.",""],
        ["2. For every AC the signal applies to, add a row in `SignalMapping` with that series_id.",""],
        ["3. Run `python src/build_dashboard.py`.",""],
        ["", ""],
        ["HOW TO ADD A NEW ASSET CLASS", ""],
        ["1. Add row in `AssetClasses` with a new ac_id (snake_case).",""],
        ["2. Add a row in `PillarWeights` for the new ac_id with F+M+S+V = 1.0.",""],
        ["3. Add SignalMapping rows for each pillar so the AC has at least one signal per pillar.",""],
        ["4. Run `python src/build_dashboard.py`.",""],
        ["", ""],
        ["CAUTION",""],
        ["src/pillars.py has hard-coded signal->AC wiring for live computation.",""],
        ["The Excel drives the DOCUMENTATION and the BLUEPRINT view in index.html.",""],
        ["When pillars.py is refactored to read SignalMapping at runtime, the Excel becomes fully live.",""],
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14, color="C41230")
    for r in (7, 14, 24, 30, 35):
        ws.cell(row=r, column=1).font = Font(bold=True, size=11, color="1E2E47")
    for r in (35,):
        ws.cell(row=r, column=1).fill = WARN_FILL
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 85
    return ws

# =============================================================================
# SHEET 1 - ASSET CLASSES
# =============================================================================
ASSET_CLASSES = [
    # id, full_label, short_label, group, benchmark, sub_description, color, max_tilt_pct
    ("money_market",  "Money Market",        "MM",     "FI", "US 3M T-Bill",                "T-bills, repo, commercial paper",               "#14B8A6", 2.0),
    ("short_term_fi", "Short-Term Fixed Income (USD)", "STFI",  "FI", "Bloomberg 1-5Y Treasury",     "1-5Y US Treasuries + IG credit",                "#3A7BD5", 3.0),
    ("lt_treasuries", "LT US Treasuries",    "LT Tsy", "FI", "Bloomberg 7-10Y Treasury",    "7-30Y duration, BSGVTRUU benchmark",            "#60A5FA", 4.0),
    ("lt_us_corp",    "LT US Corporate",     "IG Corp","FI", "Bloomberg US Corp IG",        "IG credit, 5-10Y, US Corp BBB/HY",              "#FBBF24", 3.0),
    ("lt_em_fi",      "LT EM Fixed Income",  "EM FI",  "FI", "EMBI Global Diversified",     "EMBI + EM corporate (hard currency)",           "#F97316", 3.0),
    ("us_equity",     "US Equity (Broad)",   "US Eq",  "EQ", "S&P 500 / Russell 3000",      "S&P 500 / Russell 3000 benchmark",              "#22C55E", 5.0),
    ("us_growth",     "US Growth",           "US Gro", "EQ", "S&P 500 Growth",              "Growth style tilt within US equity",            "#4ADE80", 3.0),
    ("us_value",      "US Value",            "US Val", "EQ", "S&P 500 Value",               "Value style tilt within US equity",             "#86EFAC", 3.0),
    ("dm_equity",     "DM ex-US Equity",     "DM",     "EQ", "MSCI EAFE",                   "Developed markets excluding the US",            "#3FB950", 4.0),
    ("em_equity",     "Emerging Markets Equity", "EM Eq","EQ","MSCI EM (NDUEEGF)",         "Broad emerging markets equity",                 "#C4B5FD", 4.0),
    ("em_xchina",     "EM ex-China",         "EM xCN", "EQ", "MSCI EM ex China",            "Emerging markets excluding China",              "#A78BFA", 3.0),
    ("china_equity",  "China Equity",        "China",  "EQ", "MSCI China (NDEUCHF)",        "Chinese equity (onshore + offshore)",           "#F87171", 3.0),
]

def _sheet_asset_classes(wb):
    ws = wb.create_sheet("AssetClasses")
    ws.sheet_properties.tabColor = "3A7BD5"
    headers = ["ac_id","full_label","short_label","group","benchmark","sub_description","color","max_tilt_pct"]
    ws.append(headers)
    for row in ASSET_CLASSES:
        ws.append(list(row))
    _apply_header(ws)
    _autosize(ws)
    ws.freeze_panes = "A2"
    return ws

# =============================================================================
# SHEET 2 - DATA SERIES (MASTER CATALOGUE)
# =============================================================================
# (series_id, signal_name, ticker, source, frequency, pillar, transformation, window, notes)
DATA_SERIES = [
    # --- PILLAR I: FUNDAMENTALS --------------------------------------------
    ("ism_mfg_pmi",       "ISM Mfg PMI",                   "NAPMPMI Index",      "BBG",  "Monthly", "F", "Level z-score + Delta3M + 4-quadrant regime",  "60M",   "Classic US manufacturing activity gauge"),
    ("ism_svcs_pmi",      "ISM Services PMI",              "NAPMNMI Index",      "BBG",  "Monthly", "F", "Level z-score + Delta3M",                      "60M",   "US services economy breadth"),
    ("ism_new_ord_inv",   "ISM New Orders / Inventories",  ".ISM G Index",       "BBG",  "Monthly", "F", "Ratio level z-score",                          "60M",   "Leading indicator for ISM headline"),
    ("ism_employment",    "ISM Mfg Employment",            "NAPMEMPL Index",     "BBG",  "Monthly", "F", "Level z-score",                                "60M",   "US labour demand signal"),
    ("ism_exports",       "ISM Mfg New Export Orders",     "NAPMNEWO Index",     "BBG",  "Monthly", "F", "Level z-score",                                "60M",   "Proxy for global external demand"),
    ("pmi_ez_mfg",        "Eurozone Mfg PMI",              "MPMIEZMA Index",     "BBG",  "Monthly", "F", "Level z-score + Delta3M",                      "60M",   "Eurozone factory activity"),
    ("pmi_ez_svcs",       "Eurozone Services PMI",         "MPMIEZSA Index",     "BBG",  "Monthly", "F", "Level z-score",                                "60M",   "Eurozone services activity"),
    ("pmi_china_mfg",     "China Mfg PMI (Caixin)",        "CPMINDX Index",      "BBG",  "Monthly", "F", "Level z-score + Delta3M",                      "60M",   "Dominant EM fundamental signal"),
    ("pmi_china_svcs",    "China Services PMI",            "MPMICNSA Index",     "BBG",  "Monthly", "F", "Level z-score",                                "60M",   "China consumer/services economy"),
    ("pmi_japan_mfg",     "Japan Mfg PMI",                 "MPMIJPMA Index",     "BBG",  "Monthly", "F", "Level z-score",                                "60M",   "Japanese industrial activity"),
    ("pmi_global_mfg",    "Global Mfg PMI",                "MPMIGLMA Index",     "BBG",  "Monthly", "F", "Level z-score",                                "60M",   "Cross-asset risk regime context"),
    ("cesi_us",           "Citi US Surprise (CESIUSD)",    "CESIUSD Index",      "BBG",  "Daily",   "F", "Level z + Delta20d + percentile; contrarian extremes", "252d",  "Mean-reverts to zero; flip at pctile>85 / <15"),
    ("cesi_ez",           "Citi Eurozone Surprise",        "CESIEUR Index",      "BBG",  "Daily",   "F", "Level z + Delta20d",                           "252d",  "European data surprises"),
    ("cesi_china",        "Citi China Surprise",           "CESICNY Index",      "BBG",  "Daily",   "F", "Level z + Delta20d",                           "252d",  "China macro beat/miss"),
    ("cesi_em",           "Citi EM Surprise",              "CESIEM Index",       "BBG",  "Daily",   "F", "Level z + Delta20d",                           "252d",  "Broad EM surprises"),
    ("gdp_us_rev",        "US GDP Forecast Revision",      "ECGDUS 26/27",       "BBG",  "Daily",   "F", "Blended current/next year; Delta1M revision",  "36M",   "MoM revision more predictive than level"),
    ("gdp_dm_rev",        "DM GDP Forecast Revision",      "ECGDD1 26/27",       "BBG",  "Daily",   "F", "Blended current/next; Delta1M",                "36M",   "Developed-markets consensus change"),
    ("gdp_em_rev",        "EM GDP Forecast Revision",      "ECGDM1 26/27",       "BBG",  "Daily",   "F", "Blended current/next; Delta1M",                "36M",   "EM growth forecast revisions"),
    ("gdp_china_rev",     "China GDP Forecast Revision",   "ECGDCN 26/27",       "BBG",  "Daily",   "F", "Blended current/next; Delta1M",                "36M",   "Bloomberg consensus China GDP"),
    ("gdp_eu_rev",        "EU GDP Forecast Revision",      "ECGDEU 26/27",       "BBG",  "Daily",   "F", "Blended current/next; Delta1M",                "36M",   "Eurozone GDP consensus"),
    ("eps_us_fwd",        "US Fwd EPS Growth",             "SPX BEST_EPS_GRO",   "BBG",  "Daily",   "F", "Level z + Delta20d + Earnings Revision Ratio", "252d",  "Forward EPS trend and revisions"),
    ("eps_em_fwd",        "EM Fwd EPS Growth",             "MXEF BEST_EPS_GRO",  "BBG",  "Daily",   "F", "Level z + Delta20d",                           "252d",  "MSCI EM earnings trend"),
    ("eps_ez_fwd",        "Eurozone Fwd EPS Growth",       "MXEMU BEST_EPS_GRO", "BBG",  "Daily",   "F", "Level z + Delta20d",                           "252d",  "MSCI EMU earnings trend"),
    ("eps_china_fwd",     "China Fwd EPS Growth",          "MXCN BEST_EPS_GRO",  "BBG",  "Daily",   "F", "Level z + Delta20d",                           "252d",  "MSCI China earnings trend"),
    ("eps_world_fwd",     "MSCI World Fwd EPS Growth",     "MXWO BEST_EPS_GRO",  "BBG",  "Daily",   "F", "Level z + Delta20d",                           "252d",  "DM earnings breadth"),
    ("eps_beat_us",       "US Earnings Beat %",            "Computed",           "BBG",  "Quarterly","F","Discrete score: >65% +2, 55-65% +1, <50% -1",  "60M",   "Corporate delivery vs estimates"),
    ("breakeven_5y",      "5Y Breakeven Inflation",        "T5YIE (FRED)",       "FRED", "Daily",   "F", "Level z + Delta20d vs 2% target",              "5Y",    "Market implied 5Y inflation"),
    ("breakeven_10y",     "10Y Breakeven Inflation",       "T10YIE (FRED)",      "FRED", "Daily",   "F", "Level z + Delta20d",                           "5Y",    "Amplified signal for long duration"),
    ("core_pce",          "US Core PCE YoY",               "PCEPILFE (FRED)",    "FRED", "Monthly", "F", "Level z vs 2% target",                         "60M",   "Fed's preferred inflation gauge"),
    ("fed_funds",         "Fed Funds Rate",                "FDTR (BBG)",         "BBG",  "Daily",   "F", "Level vs neutral r*; direction; real FF",      "5Y",    "Policy stance; restrictive >r*"),

    # --- PILLAR II: MOMENTUM -----------------------------------------------
    ("spxt_mom",          "SPXT 12-1M Price Momentum",     "SPXT Index TR",      "BBG",  "Daily",   "M", "12-1M(40%)+3M(25%)+MA50/200(20%)+RSI(15%)",    "1Y",    "S&P 500 composite momentum; highest IC single signal"),
    ("spxt_ma",           "SPXT MA50 vs MA200",            "SPXT Index TR",      "BBG",  "Daily",   "M", "Distance z-score",                             "1Y",    "Golden/death cross proximity"),
    ("spxt_rsi",          "RSI(14) on SPXT",               "SPXT Index TR",      "BBG",  "Daily",   "M", "(RSI-50)/50; contrarian at extremes",          "60d",   "Short-term mean-reversion"),
    ("eafe_mom",          "EAFE Price Momentum",           "M0EFHUSD Index TR",  "BBG",  "Daily",   "M", "Composite (12-1M/3M/MA/RSI)",                  "1Y",    "EAFE (DM ex-US) trend"),
    ("em_mom",            "MSCI EM Price Momentum",        "NDUEEGF Index TR",   "BBG",  "Daily",   "M", "Composite",                                    "1Y",    "Broad EM equity trend"),
    ("em_xcn_mom",        "EM ex-China Price Momentum",    "M1CXBRV Index TR",   "BBG",  "Daily",   "M", "Composite",                                    "1Y",    "EM ex-China breadth check"),
    ("china_mom",         "China Price Momentum",          "NDEUCHF Index TR",   "BBG",  "Daily",   "M", "Composite",                                    "1Y",    "MSCI China trend"),
    ("growth_mom",        "US Growth Momentum",            "SPTRSGX Index TR",   "BBG",  "Daily",   "M", "Composite",                                    "1Y",    "S&P 500 Growth trend"),
    ("value_mom",         "US Value Momentum",             "SPTRSVX Index TR",   "BBG",  "Daily",   "M", "Composite",                                    "1Y",    "S&P 500 Value trend"),
    ("bfu5_mom",          "BFU5TRUU 1-5Y Treasury Mom",    "BFU5TRUU Index",     "BBG",  "Daily",   "M", "Composite, TR-based",                          "1Y",    "Short UST price momentum"),
    ("i132_mom",          "ICE 1-3Y UST Momentum",         "I13282US Index",     "BBG",  "Daily",   "M", "Composite, TR-based",                          "1Y",    "Short UST momentum (alt)"),
    ("bsgv_mom",          "BSGVTRUU Long Govt Mom",        "BSGVTRUU Index",     "BBG",  "Daily",   "M", "Composite, TR-based",                          "1Y",    "Long UST TR trend"),
    ("usagg_mom",         "US Agg Momentum",               "I26729US Index",     "BBG",  "Daily",   "M", "Composite",                                    "1Y",    "IG Corp proxy"),
    ("lt03_mom",          "LT03TRUU ST Govt Mom",          "LT03TRUU Index",     "BBG",  "Daily",   "M", "Composite",                                    "1Y",    "Money market TR proxy"),
    ("bbb_oas_mom",       "BBB OAS Momentum",              "BAMLC0A4CBBB",       "FRED", "Daily",   "M", "z = -1 * zscore(DeltaOAS)",                    "5Y",    "Tightening=+ for IG credit"),
    ("hy_oas_mom",        "HY OAS Momentum",               "BAMLH0A0HYM2",       "FRED", "Daily",   "M", "z = -1 * zscore(DeltaOAS)",                    "5Y",    "Tightening=+ for HY / equity proxy"),
    ("em_oas_mom",        "EM BBB OAS Momentum",           "BAMLEM2BRRBBBCRPIOAS","FRED","Daily",   "M", "z = -1 * zscore(DeltaOAS)",                    "5Y",    "EM credit leading indicator"),
    ("latam_oas_mom",     "LatAm Corp OAS Momentum",       "BAMLEMRLCRPILAOAS",  "FRED", "Daily",   "M", "z = -1 * zscore(DeltaOAS)",                    "5Y",    "LatAm-specific spread mom"),
    ("cdx_ig_mom",        "CDX IG 5Y Momentum",            "IBOXUMAE Index",     "BBG",  "Daily",   "M", "z = -1 * zscore(DeltaSpread)",                 "5Y",    "Real-time IG credit conditions"),
    ("cdx_hy_mom",        "CDX HY Momentum",               "IBOXHYAE Index",     "BBG",  "Daily",   "M", "Price-based composite",                        "5Y",    "HY risk appetite"),
    ("ust10_yld_mom",     "US 10Y Yield Momentum",         "GT10 @BGN Govt",     "BBG",  "Daily",   "M", "z = -1 * zscore(Deltayld)",                    "5Y",    "Falling 10Y = +z for duration"),
    ("ust2_yld_mom",      "US 2Y Yield Momentum",          "GT02 @BGN Govt",     "BBG",  "Daily",   "M", "z = -1 * zscore(Deltayld)",                    "5Y",    "Falling 2Y = +z for STFI"),
    ("dxy_mom",           "USD DXY Momentum",              "DXY Index",          "BBG",  "Daily",   "M", "Composite (direction)",                        "5Y",    "Weak USD = + for EM"),

    # --- PILLAR III: SENTIMENT ---------------------------------------------
    ("vix_level",         "VIX Level (percentile)",        "VIXCLS (FRED)",      "FRED", "Daily",   "S", "Percentile (5Y) + non-linear score",           "5Y",    "Contrarian for equity; flight-to-quality for FI"),
    ("vix_term",          "VIX Term Structure (VIX/VIX3M)","Computed",           "FRED", "Daily",   "S", "Ratio; >1 = backwardation (panic)",            "5Y",    "Near-term stress vs medium-term"),
    ("move_index",        "MOVE Index",                    "MOVE Index",         "BBG",  "Daily",   "S", "Level z; regime thresholds",                   "5Y",    "Bond vol; high = bad for FI carry"),
    ("vstoxx",            "VSTOXX (EZ vol)",               "V2X Index",          "BBG",  "Daily",   "S", "Percentile (5Y)",                              "5Y",    "DM ex-US vol"),
    ("ted_spread",        "TED Spread",                    "TEDRATE (FRED)",     "FRED", "Daily",   "S", "Level z (5Y)",                                 "5Y",    "Funding stress; risk-off gauge"),
    ("dxy_level",         "USD DXY Level",                 "DTWEXBGS (FRED)",    "FRED", "Daily",   "S", "Level z inverted for EM",                      "5Y",    "Strong USD = EM headwind"),
    ("embi_spread",       "EMBI Sovereign Spread",         "BAMLHE00EHY2Y",      "FRED", "Daily",   "S", "Level z; widening = negative",                 "5Y",    "EM sovereign stress"),
    ("gsfci",             "GS Financial Conditions",       "GSFCI Index",        "BBG",  "Daily",   "S", "Level z + Delta20d",                           "5Y",    "Broad US financial conditions"),
    ("pcr_10d",           "CBOE PCR 10d Avg",              "CPCE (FRED)",        "FRED", "Daily",   "S", "MA10 level z; contrarian",                     "5Y",    ">1.1 fear -> buy; <0.7 -> sell"),
    ("aaii_bb",           "AAII Bull-Bear Spread",         "AAII.com",           "AAII", "Weekly",  "S", "Spread z (104w); contrarian",                  "2Y",    ">+30% sell; <-20% buy"),
    ("cot_spx",           "CFTC COT S&P Net Spec",         "CFTC / Quandl",      "CFTC", "Weekly",  "S", "Net positions z; contrarian",                  "1Y",    "Extreme long = crowded = fade"),
    ("cot_ust10",         "CFTC COT UST 10Y Net Spec",     "CFTC / Quandl",      "CFTC", "Weekly",  "S", "Net positions z; contrarian",                  "1Y",    "Extreme short = squeeze = buy UST"),
    ("skew",              "SKEW Index",                    "SKEWX (FRED)",       "FRED", "Daily",   "S", "Level z (252d)",                               "1Y",    "Tail hedging; mild equity bearish"),

    # --- PILLAR IV: VALUATION ----------------------------------------------
    ("pe_spx",            "S&P 500 Forward P/E",           "SPXT Index PE",      "BBG",  "Daily",   "V", "pctile(10Y); score = -2 + 4*(1 - pctile)",     "10Y",   "<15x cheap / 15-20x fair / >22x expensive"),
    ("pe_eafe",           "EAFE Forward P/E",              "M0EFHUSD Index PE",  "BBG",  "Daily",   "V", "pctile(10Y)",                                  "10Y",   "<13x cheap / 13-17x fair / >19x expensive"),
    ("pe_em",             "MSCI EM Forward P/E",           "NDUEEGF Index PE",   "BBG",  "Daily",   "V", "pctile(10Y)",                                  "10Y",   "<10x cheap / 10-13x fair / >15x expensive"),
    ("pe_china",          "MSCI China Forward P/E",        "NDEUCHF Index PE",   "BBG",  "Daily",   "V", "pctile(10Y)",                                  "10Y",   "<9x cheap / 9-12x fair / >14x expensive"),
    ("pe_em_xcn",         "MSCI EM ex-China Forward P/E",  "M1CXBRV Index PE",   "BBG",  "Daily",   "V", "pctile(10Y)",                                  "10Y",   "<10x cheap / 10-13x fair"),
    ("pe_gro",            "S&P 500 Growth P/E",            "SPTRSGX Index PE",   "BBG",  "Daily",   "V", "Relative to Value: z-score of ratio",          "10Y",   "Growth/Value relative valuation"),
    ("pe_val",            "S&P 500 Value P/E",             "SPTRSVX Index PE",   "BBG",  "Daily",   "V", "Relative to Market: z-score of ratio",         "10Y",   "Value style attractiveness"),
    ("erp_us",            "US Equity Risk Premium",        "Computed (EY-TIPS10)","BBG", "Daily",   "V", "z-score (10Y)",                                "10Y",   "Primary cross-asset signal; >4% = OW equity"),
    ("erp_em",            "EM Equity Risk Premium",        "Computed",           "BBG",  "Daily",   "V", "z-score (10Y)",                                "10Y",   "EM EY - UST 10Y - EMBI spread"),
    ("shiller_cape",      "Shiller CAPE (10Y)",            "Computed",           "BBG",  "Daily",   "V", "Level z; >35x = expensive",                    "30Y",   "Long-run anchor (10% weight cap)"),
    ("yield_ust10",       "US 10Y Yield Level (pctile)",   "GT10 @BGN Govt",     "BBG",  "Daily",   "V", "pctile(10Y)",                                  "10Y",   "High yield = attractive UST carry"),
    ("yield_ust2",        "US 2Y Yield Level (pctile)",    "GT02 @BGN Govt",     "BBG",  "Daily",   "V", "pctile(10Y)",                                  "10Y",   "High 2Y = attractive STFI carry"),
    ("term_spread",       "Term Spread 10Y-2Y",            "Computed",           "BBG",  "Daily",   "V", "Level z; inverted = bullish duration",         "10Y",   "Recession signal when inverted"),
    ("tips10",            "TIPS 10Y Real Yield",           "DFII10 (FRED)",      "FRED", "Daily",   "V", "Level z",                                      "10Y",   ">2% = duration very attractive"),
    ("tips5",             "TIPS 5Y Real Yield",            "DFII5 (FRED)",       "FRED", "Daily",   "V", "Level z",                                      "10Y",   ">1.5% = STFI very attractive"),
    ("bbb_oas_lv",        "BBB OAS Level (pctile)",        "BAMLC0A4CBBB",       "FRED", "Daily",   "V", "pctile(5Y/10Y); >180bps cheap",                "10Y",   "IG valuation anchor"),
    ("hy_oas_lv",         "HY OAS Level (pctile)",         "BAMLH0A0HYM2",       "FRED", "Daily",   "V", "pctile; >500bps cheap",                        "10Y",   "HY carry vs risk"),
    ("em_oas_lv",         "EM BBB OAS Level (pctile)",     "BAMLEM2BRRBBBCRPIOAS","FRED","Daily",   "V", "pctile; >250bps cheap",                        "10Y",   "EM credit cheapness"),
    ("latam_oas_lv",      "LatAm OAS Level",               "BAMLEMRLCRPILAOAS",  "FRED", "Daily",   "V", "pctile",                                       "10Y",   "LatAm-specific credit"),
    ("real_tbill_3m",     "Real 3M T-bill Rate",           "Computed (DTB3-CPI)","FRED", "Daily",   "V", "Level z",                                      "10Y",   "Main MM signal"),
    ("tbill_3m_pctile",   "3M T-bill Yield Pctile",        "DTB3 (FRED)",        "FRED", "Daily",   "V", "pctile(10Y)",                                  "10Y",   "Nominal cash yield attractiveness"),
    ("ff_vs_neutral",     "Fed Funds vs Neutral (r*)",     "FEDFUNDS (FRED)",    "FRED", "Daily",   "V", "Level vs 2.5% neutral",                        "10Y",   "Policy stance"),
    ("hy_ig_ratio",       "HY/IG Spread Ratio",            "Computed",           "FRED", "Daily",   "V", "Level z of ratio",                             "10Y",   "Credit cycle position"),
]

def _sheet_data_series(wb):
    ws = wb.create_sheet("DataSeries")
    ws.sheet_properties.tabColor = "14B8A6"
    headers = ["series_id","signal_name","ticker","source","frequency","pillar","transformation","window","notes"]
    ws.append(headers)
    for row in DATA_SERIES:
        ws.append(list(row))
    _apply_header(ws)
    _autosize(ws, max_w=60)
    ws.freeze_panes = "B2"
    return ws

# =============================================================================
# SHEET 3 - PILLAR WEIGHTS
# =============================================================================
PILLAR_WEIGHTS = {
    "money_market":  (0.10, 0.15, 0.25, 0.50),
    "short_term_fi": (0.20, 0.25, 0.20, 0.35),
    "lt_treasuries": (0.25, 0.25, 0.20, 0.30),
    "lt_us_corp":    (0.20, 0.30, 0.20, 0.30),
    "lt_em_fi":      (0.25, 0.30, 0.20, 0.25),
    "us_equity":     (0.25, 0.30, 0.20, 0.25),
    "us_growth":     (0.20, 0.35, 0.15, 0.30),
    "us_value":      (0.30, 0.25, 0.20, 0.25),
    "dm_equity":     (0.25, 0.30, 0.20, 0.25),
    "em_equity":     (0.25, 0.30, 0.20, 0.25),
    "em_xchina":     (0.25, 0.30, 0.20, 0.25),
    "china_equity":  (0.25, 0.30, 0.20, 0.25),
}

def _sheet_pillar_weights(wb):
    ws = wb.create_sheet("PillarWeights")
    ws.sheet_properties.tabColor = "F59E0B"
    ws.append(["ac_id","F","M","S","V","total_check"])
    for ac, (f,m,s,v) in PILLAR_WEIGHTS.items():
        ws.append([ac, f, m, s, v, f + m + s + v])
    _apply_header(ws)
    _autosize(ws)
    ws.freeze_panes = "A2"
    return ws

# =============================================================================
# SHEET 4 - PILLAR NOTES (optional italic header per AC+pillar)
# =============================================================================
PILLAR_NOTES = [
    # (ac_id, pillar, note)
    ("money_market",  "F", "Macro is mostly context here - primarily used for regime detection, not direct signal."),
    ("money_market",  "M", "MM instruments don't have trending returns, so momentum is limited."),
    ("money_market",  "S", "Flight to safety is the primary driver of MM demand."),
    ("money_market",  "V", "The dominant signal for MM is simply: is the real short-term rate positive and attractive?"),
    ("short_term_fi", "F", "Growth/PMI signals are inverted: strong growth -> rates rise -> STFI price falls. Credit fundamentals remain direct (growth = tighter spreads)."),
    ("short_term_fi", "M", "Use TR price directly. Falling prices = negative momentum. OAS tightening on short-end credits is additional carry signal."),
    ("short_term_fi", "S", "Flight to quality (VIX, MOVE, TED) is positive for UST/STFI. USD strength is neutral for STFI."),
    ("short_term_fi", "V", "Carry is the primary valuation signal. 2Y yield level and real yield determine attractiveness."),
    ("lt_treasuries", "F", "All macro growth signals are inverted and amplified. Inflation signals are doubly negative (higher rates + lower real value)."),
    ("lt_treasuries", "M", "Long UST have significant price momentum. Duration sells off in trending rate-rise environments and rallies in risk-off."),
    ("lt_treasuries", "S", "VIX and MOVE are positive for UST. CFTC net-short UST (contrarian) also matters."),
    ("lt_treasuries", "V", "10Y yield level, real yield, and term premium are the anchors. Most powerful: high real yields + inverted curve."),
    ("lt_us_corp",    "F", "Dual nature: duration (rates) + spread (credit). Spread effect typically dominates IG over short run."),
    ("lt_us_corp",    "M", "OAS tightening momentum is primary. Price momentum of TR index is secondary."),
    ("lt_us_corp",    "S", "VIX mildly negative for IG spreads. Flight to quality from HY benefits IG."),
    ("lt_us_corp",    "V", "OAS level is the primary valuation signal. BBB OAS >180bps = cheap; <100bps = rich."),
    ("lt_em_fi",      "F", "Driven by EM growth, China activity, and commodity cycles. Growth signals positive because spread component dominates duration."),
    ("lt_em_fi",      "M", "EM OAS tightening + EM equity momentum (leading) + USD weakness momentum."),
    ("lt_em_fi",      "S", "USD strength is the #1 headwind. EMBI level proxies sovereign risk."),
    ("lt_em_fi",      "V", "Spread level (EMBI + EM corp OAS) is the primary anchor. EM real yield vs DM is a key RV signal."),
    ("us_equity",     "F", "US growth signals are direct. ISM PMI + CESI + GDP revisions + EPS growth drive the composite."),
    ("us_equity",     "M", "12-1M price momentum is the single highest-IC signal. Credit momentum (HY) leads by 2-4 weeks."),
    ("us_equity",     "S", "All sentiment signals are contrarian at extremes. VIX >80%ile = buy; <20%ile = sell."),
    ("us_equity",     "V", "ERP (EY - TIPS 10Y) is the primary cross-asset valuation signal. P/E percentile is secondary."),
    ("dm_equity",     "F", "Eurozone PMI dominates (EAFE ~60% Europe). Japan is secondary."),
    ("dm_equity",     "M", "EAFE TR momentum. USD-hedged removes currency noise."),
    ("dm_equity",     "S", "VIX + VSTOXX for global risk-off. USD weakness = better USD-terms returns."),
    ("dm_equity",     "V", "EAFE P/E percentile + ERP + relative vs US."),
    ("em_equity",     "F", "China PMI is dominant (~30%+ of MSCI EM). Broad EM surprises + EM GDP revisions."),
    ("em_equity",     "M", "MSCI EM TR + EM ex-China breadth check + EM credit leads by 2-4 weeks."),
    ("em_equity",     "S", "USD strength = #1 headwind. EMBI spread proxies sovereign risk."),
    ("em_equity",     "V", "P/E percentile + EM ERP + EM credit cheapness."),
    ("china_equity",  "F", "Caixin PMI dominant. Level >50 + rising direction = maximum bullish."),
    ("china_equity",  "M", "MSCI China often diverges from broad EM. Monitor both absolute and China vs EM xCN."),
    ("china_equity",  "S", "CNY weakness = capital flight signal. EMBI / China sovereign spreads matter."),
    ("china_equity",  "V", "China often <10x P/E = cheap; <8x = very cheap. Discount to US narrowing = less attractive."),
    ("us_growth",     "F", "Inherits US equity signals with stronger EPS-growth tilt."),
    ("us_growth",     "M", "S&P 500 Growth TR composite momentum; HY credit support."),
    ("us_growth",     "S", "Same as US equity; contrarian sentiment."),
    ("us_growth",     "V", "Growth vs Value relative P/E is the key style signal."),
    ("us_value",      "F", "Emphasises cyclical signals (PMI, GDP). Earnings beats supportive."),
    ("us_value",      "M", "SPTRSVX TR + IG OAS (value benefits from IG tightening)."),
    ("us_value",      "S", "US equity contrarian sentiment."),
    ("us_value",      "V", "Value vs Growth relative P/E + market P/E."),
    ("em_xchina",     "F", "Excludes China-specific drivers; emphasises broader EM PMI/CESI/EPS."),
    ("em_xchina",     "M", "MSCI EM ex-China TR composite."),
    ("em_xchina",     "S", "Same EM sentiment signals as broad EM."),
    ("em_xchina",     "V", "MSCI EM ex-China P/E percentile + EM credit level."),
]

def _sheet_pillar_notes(wb):
    ws = wb.create_sheet("PillarNotes")
    ws.sheet_properties.tabColor = "A855F7"
    ws.append(["ac_id","pillar","note"])
    for row in PILLAR_NOTES:
        ws.append(list(row))
    _apply_header(ws)
    _autosize(ws, max_w=110)
    ws.freeze_panes = "A2"
    return ws

# =============================================================================
# SHEET 5 - SIGNAL MAPPING
# =============================================================================
# Sign tokens: ++, +, -, --, C (contrarian), n/a (not applicable)
# weight_in_pillar is a percent string so it renders cleanly in the dashboard.
# (ac_id, series_id, pillar, sign, weight_in_pillar, description_override)
SIGNAL_MAPPING = [
    # =========================================================================
    # MONEY MARKET
    # =========================================================================
    ("money_market", "ff_vs_neutral",  "F", "+", "40%", "High real FF = cash yield attractive; FF above neutral = restrictive = MM rates stay high"),
    ("money_market", "core_pce",       "F", "+", "30%", "Inflation above target = Fed holds rates higher = positive for MM carry"),
    ("money_market", "breakeven_5y",   "F", "+", "30%", "Rising breakeven = Fed can't cut -> positive for MM rates"),
    ("money_market", "ust2_yld_mom",   "M", "+", "60%", "Rising 3M/2Y yield = positive momentum for cash"),
    ("money_market", "lt03_mom",       "M", "+", "40%", "Secondary FI TR signal for positioning"),
    ("money_market", "ted_spread",     "S", "+", "35%", "High TED = funding stress = demand for cash/safety -> OW MM"),
    ("money_market", "vix_level",      "S", "+", "35%", "High VIX = risk-off = demand for cash equivalents"),
    ("money_market", "move_index",     "S", "+", "30%", "High bond vol = prefer cash over duration"),
    ("money_market", "real_tbill_3m",  "V", "+", "40%", "3M yield - CPI. Positive and high = MM very attractive. Main signal."),
    ("money_market", "tbill_3m_pctile","V", "+", "30%", "Is the nominal yield high vs history?"),
    ("money_market", "ff_vs_neutral",  "V", "+", "30%", "FF > r* (2.5%) = policy restrictive = MM attractive vs duration"),

    # =========================================================================
    # SHORT-TERM FI (INVERTED for growth)
    # =========================================================================
    ("short_term_fi","ism_mfg_pmi",    "F", "-", "30%", "PMI accelerating -> rates rise -> STFI price falls. Sign: -1 * z_pmi"),
    ("short_term_fi","cesi_us",        "F", "-", "25%", "Positive surprises -> rate repricing upward -> -1 * z_cesi"),
    ("short_term_fi","gdp_us_rev",     "F", "-", "25%", "Upward GDP revision -> rates rise. Sign: -1 * z_gdp_revision"),
    ("short_term_fi","breakeven_5y",   "F", "-", "20%", "Rising inflation expectations -> nominal rates up -> -1 * z_breakeven"),
    ("short_term_fi","bfu5_mom",       "M", "+", "45%", "12-1M momentum most predictive; weighted avg of horizons"),
    ("short_term_fi","ust2_yld_mom",   "M", "+", "25%", "z_yield_mom = -1 * zscore(Delta2Y_yield_1M)"),
    ("short_term_fi","bbb_oas_mom",    "M", "+", "30%", "z_oas = -1 * zscore(DeltaBAMLC0A4CBBB, 1M)"),
    ("short_term_fi","vix_level",      "S", "+", "35%", "High VIX = demand for safe assets -> OW STFI"),
    ("short_term_fi","move_index",     "S", "C", "35%", "MOVE > 120: reduce duration risk; below 80: favourable carry"),
    ("short_term_fi","ted_spread",     "S", "+", "30%", "High TED = demand for T-bills and short UST"),
    ("short_term_fi","yield_ust2",     "V", "+", "35%", "High 2Y yield = attractive carry. Main valuation signal."),
    ("short_term_fi","tips5",          "V", "+", "25%", "Real yield > 1.5% = very attractive; < 0% = unattractive"),
    ("short_term_fi","bbb_oas_lv",     "V", "+", "20%", "High OAS = credit component cheap = higher total yield"),
    ("short_term_fi","term_spread",    "V", "+", "20%", "Inverted: high short-end yield vs long end = favourable STFI carry"),

    # =========================================================================
    # LT TREASURIES (STRONGLY INVERTED)
    # =========================================================================
    ("lt_treasuries","ism_mfg_pmi",    "F", "-",  "30%", "PMI Quadrant 3/4 (slowdown/contraction) -> bullish duration"),
    ("lt_treasuries","gdp_us_rev",     "F", "-",  "25%", "Downward revisions -> rate cut pricing -> duration rally"),
    ("lt_treasuries","breakeven_10y",  "F", "--", "25%", "Rising breakevens = most bearish signal for nominal UST"),
    ("lt_treasuries","cesi_us",        "F", "-",  "20%", "Negative surprises -> rate cut pricing -> duration bullish"),
    ("lt_treasuries","bsgv_mom",       "M", "+",  "45%", "12-1M skip: most predictive for duration assets"),
    ("lt_treasuries","ust10_yld_mom",  "M", "+",  "30%", "z = -1 * zscore(Delta10Y_yield_1M/3M)"),
    ("lt_treasuries","bbb_oas_mom",    "M", "+",  "25%", "When credit widens rapidly -> flight to quality = UST rally"),
    ("lt_treasuries","vix_level",      "S", "+",  "35%", "Crisis signal: VIX > 80th pctile = strong buy UST"),
    ("lt_treasuries","cot_ust10",      "S", "C",  "35%", "Extreme net short = crowded = potential squeeze = buy UST"),
    ("lt_treasuries","move_index",     "S", "C",  "30%", "MOVE > 150: override to neutral. MOVE < 70: carry environment"),
    ("lt_treasuries","tips10",         "V", "+",  "40%", "Real yield > 2.0% = duration very attractive. Primary signal."),
    ("lt_treasuries","yield_ust10",    "V", "+",  "30%", "High pctile = high nominal yield = better carry"),
    ("lt_treasuries","term_spread",    "V", "+",  "30%", "Deeply inverted curve = recession/rate-cut signal = OW duration"),

    # =========================================================================
    # LT US CORPORATE
    # =========================================================================
    ("lt_us_corp",   "eps_us_fwd",     "F", "+", "30%", "Strong earnings = tighter spreads = positive for IG credit"),
    ("lt_us_corp",   "ism_mfg_pmi",    "F", "+", "25%", "Expansion = tighter spreads. But rising rates offset. Net mild positive."),
    ("lt_us_corp",   "eps_beat_us",    "F", "+", "25%", "High beat rate = credit quality improving = tighter spreads"),
    ("lt_us_corp",   "gdp_us_rev",     "F", "+", "20%", "Upward revision = stronger balance sheets = tighter spreads"),
    ("lt_us_corp",   "usagg_mom",      "M", "+", "30%", "12-1M, 3M, 1M returns. Most predictive horizon: 3M."),
    ("lt_us_corp",   "bbb_oas_mom",    "M", "+", "35%", "z = -1 * zscore(Delta OAS). Primary spread momentum signal."),
    ("lt_us_corp",   "cdx_ig_mom",     "M", "+", "20%", "IG CDS market as real-time credit conditions gauge"),
    ("lt_us_corp",   "ust10_yld_mom",  "M", "+", "15%", "Duration component: falling 10Y = positive for IG price"),
    ("lt_us_corp",   "vix_level",      "S", "-", "35%", "Rising VIX -> credit spread widening. But IG benefits vs HY."),
    ("lt_us_corp",   "gsfci",          "S", "-", "35%", "Tighter conditions -> wider spreads -> negative for IG"),
    ("lt_us_corp",   "cdx_hy_mom",     "S", "+", "30%", "HY outperforming IG = risk-on = tighter IG spreads"),
    ("lt_us_corp",   "bbb_oas_lv",     "V", "+", "40%", "Main valuation anchor for IG corporate spread component"),
    ("lt_us_corp",   "hy_ig_ratio",    "V", "+", "25%", "Ratio > 4x = HY expensive, rotate to IG = IG attractive relatively"),
    ("lt_us_corp",   "tips10",         "V", "+", "20%", "Duration component valuation: real yield attractiveness"),
    ("lt_us_corp",   "hy_oas_lv",      "V", "+", "15%", "HY level as credit cycle regime"),

    # =========================================================================
    # LT EM FIXED INCOME
    # =========================================================================
    ("lt_em_fi",     "gdp_em_rev",     "F", "+", "30%", "Upward EM GDP revision = tighter EMBI spreads = positive"),
    ("lt_em_fi",     "pmi_china_mfg",  "F", "+", "25%", "China is engine of EM growth. PMI expansion = positive for EM FI"),
    ("lt_em_fi",     "eps_em_fwd",     "F", "+", "25%", "Corporate fundamentals of EM issuers"),
    ("lt_em_fi",     "cesi_em",        "F", "+", "20%", "Positive EM surprises = risk appetite for EM debt"),
    ("lt_em_fi",     "em_oas_mom",     "M", "+", "30%", "Primary spread momentum. z = -1 * zscore(DeltaOAS)"),
    ("lt_em_fi",     "latam_oas_mom",  "M", "+", "20%", "LatAm-specific spread momentum"),
    ("lt_em_fi",     "em_mom",         "M", "+", "25%", "EM equity often leads EM credit by 1-3 months"),
    ("lt_em_fi",     "dxy_mom",        "M", "+", "25%", "Dollar weakening = EM assets rally = EM FI spread tightens"),
    ("lt_em_fi",     "dxy_level",      "S", "-", "30%", "Strong USD = EM capital outflows = EM FI negative. z = -1 * z_DXY"),
    ("lt_em_fi",     "embi_spread",    "S", "-", "30%", "Widening EMBI = systemic EM sovereign stress = negative"),
    ("lt_em_fi",     "vix_level",      "S", "-", "20%", "High VIX = exit EM assets = negative"),
    ("lt_em_fi",     "em_oas_lv",      "S", "+", "20%", "Stable EM credit = positive sentiment signal"),
    ("lt_em_fi",     "em_oas_lv",      "V", "+", "35%", "High OAS = cheap EM credit. > 250 bps = attractive."),
    ("lt_em_fi",     "latam_oas_lv",   "V", "+", "20%", "LatAm-specific credit cheapness"),
    ("lt_em_fi",     "tips10",         "V", "+", "25%", "EM yields - DM yields - EMBI sovereign spread = EM excess real yield"),
    ("lt_em_fi",     "breakeven_10y",  "V", "-", "20%", "US inflation = USD appreciation risk = negative for USD-denominated EM"),

    # =========================================================================
    # US EQUITY (broad)
    # =========================================================================
    ("us_equity",    "ism_mfg_pmi",    "F", "+", "35%", "PMI 4-quadrant regime. Quadrant 1/2 -> OW equity."),
    ("us_equity",    "cesi_us",        "F", "+", "25%", "Positive surprises (within 15-85 pctile) = bullish. Contrarian at extremes."),
    ("us_equity",    "gdp_us_rev",     "F", "+", "20%", "Consensus GDP rising = earnings forecast uplift"),
    ("us_equity",    "eps_us_fwd",     "F", "+", "15%", "EPS growth > 10% and rising = high conviction bull signal"),
    ("us_equity",    "eps_beat_us",    "F", "+", "5%",  ">65% beat rate = strong corporate delivery"),
    ("us_equity",    "spxt_mom",       "M", "+", "40%", "Cross-sectional + time-series. Highest IC single signal."),
    ("us_equity",    "spxt_ma",        "M", "+", "25%", "Golden cross = +1. Distance z-score = trend confirmation."),
    ("us_equity",    "spxt_rsi",       "M", "C", "20%", "Continuous signal (RSI-50)/50. Contrarian at extremes >80/<20."),
    ("us_equity",    "hy_oas_mom",     "M", "+", "15%", "Credit conditions leading equity by ~2-4 weeks"),
    ("us_equity",    "vix_level",      "S", "C", "35%", "VIX > 80%ile = buy signal. VIX < 20%ile = sell signal."),
    ("us_equity",    "pcr_10d",        "S", "C", "25%", "PCR > 1.1 = fear = contrarian buy. PCR < 0.7 = sell."),
    ("us_equity",    "aaii_bb",        "S", "C", "20%", "+30% net bulls = sell. -20% net bulls = buy."),
    ("us_equity",    "cot_spx",        "S", "C", "20%", "Extreme net long = crowded = fade"),
    ("us_equity",    "pe_spx",         "V", "+", "35%", "pctile < 20% = cheap. pctile > 80% = expensive. Score: -2 + 4*(1-pctile)"),
    ("us_equity",    "erp_us",         "V", "+", "35%", "ERP > 4% = strongly OW equity vs bonds. Primary cross-asset signal."),
    ("us_equity",    "pe_eafe",        "V", "-", "20%", "US P/E / EAFE P/E z-score. High ratio = US expensive = reduce tilt"),
    ("us_equity",    "shiller_cape",   "V", "+", "10%", "Long-run anchor. > 35x = expensive territory. Use as 10% weight cap."),

    # =========================================================================
    # US GROWTH
    # =========================================================================
    ("us_growth",    "ism_mfg_pmi",    "F", "+", "30%", "Growth sectors benefit from expansion phase"),
    ("us_growth",    "eps_us_fwd",     "F", "+", "30%", "EPS acceleration = growth tailwind"),
    ("us_growth",    "cesi_us",        "F", "+", "25%", "Surprise regime"),
    ("us_growth",    "gdp_us_rev",     "F", "+", "15%", "Consensus GDP revisions"),
    ("us_growth",    "growth_mom",     "M", "+", "55%", "SPTRSGX TR composite momentum"),
    ("us_growth",    "spxt_mom",       "M", "+", "25%", "Broad US equity momentum"),
    ("us_growth",    "hy_oas_mom",     "M", "+", "20%", "HY credit risk appetite supports growth"),
    ("us_growth",    "vix_level",      "S", "C", "40%", "Contrarian at extremes"),
    ("us_growth",    "pcr_10d",        "S", "C", "30%", "Options positioning contrarian"),
    ("us_growth",    "aaii_bb",        "S", "C", "30%", "Retail sentiment contrarian"),
    ("us_growth",    "pe_gro",         "V", "+", "40%", "Relative to Value: z-score of ratio"),
    ("us_growth",    "erp_us",         "V", "+", "30%", "ERP applied to growth side"),
    ("us_growth",    "pe_spx",         "V", "+", "20%", "Absolute growth P/E percentile"),
    ("us_growth",    "shiller_cape",   "V", "+", "10%", "Long-run valuation anchor"),

    # =========================================================================
    # US VALUE
    # =========================================================================
    ("us_value",     "ism_mfg_pmi",    "F", "+", "35%", "Cyclical sectors drive value"),
    ("us_value",     "gdp_us_rev",     "F", "+", "25%", "GDP upgrades support cyclicals"),
    ("us_value",     "eps_beat_us",    "F", "+", "20%", "Earnings delivery supports value re-rating"),
    ("us_value",     "cesi_us",        "F", "+", "20%", "Surprise momentum"),
    ("us_value",     "value_mom",      "M", "+", "55%", "SPTRSVX TR composite"),
    ("us_value",     "bbb_oas_mom",    "M", "+", "25%", "Value benefits from IG tightening"),
    ("us_value",     "cdx_ig_mom",     "M", "+", "20%", "IG credit conditions"),
    ("us_value",     "vix_level",      "S", "C", "45%", "Contrarian at extremes"),
    ("us_value",     "pcr_10d",        "S", "C", "30%", "Options positioning"),
    ("us_value",     "aaii_bb",        "S", "C", "25%", "Sentiment contrarian"),
    ("us_value",     "pe_val",         "V", "+", "40%", "Value style attractiveness"),
    ("us_value",     "erp_us",         "V", "+", "30%", "ERP applied to value side"),
    ("us_value",     "pe_spx",         "V", "+", "20%", "Overall market P/E level"),
    ("us_value",     "shiller_cape",   "V", "+", "10%", "Long-run anchor"),

    # =========================================================================
    # DM ex-US EQUITY
    # =========================================================================
    ("dm_equity",    "pmi_ez_mfg",     "F", "+", "30%", "Eurozone is ~60% of EAFE. PMI regime drives DM ex-US earnings"),
    ("dm_equity",    "cesi_ez",        "F", "+", "25%", "European data surprises = direct signal for Eurozone/EAFE equity"),
    ("dm_equity",    "pmi_japan_mfg",  "F", "+", "15%", "Japan ~20% of EAFE. Secondary signal."),
    ("dm_equity",    "gdp_dm_rev",     "F", "+", "20%", "DM growth consensus rising = positive for EAFE earnings"),
    ("dm_equity",    "eps_world_fwd",  "F", "+", "10%", "DM earnings breadth and trend"),
    ("dm_equity",    "eafe_mom",       "M", "+", "40%", "EAFE price momentum. Primary signal. USD-hedged removes currency noise."),
    ("dm_equity",    "spxt_ma",        "M", "+", "25%", "Trend confirmation via moving average spread"),
    ("dm_equity",    "spxt_mom",       "M", "+", "20%", "DM vs US relative return (3M/6M)"),
    ("dm_equity",    "eafe_mom",       "M", "C", "15%", "Contrarian at extremes"),
    ("dm_equity",    "vix_level",      "S", "C", "40%", "Both high = global risk-off = UW DM ex-US equity"),
    ("dm_equity",    "vstoxx",         "S", "C", "30%", "EZ-specific vol; USD weakness = DM returns better in USD terms"),
    ("dm_equity",    "dxy_level",      "S", "-", "30%", "European ETF fund flows (proxy via DXY direction)"),
    ("dm_equity",    "pe_eafe",        "V", "+", "35%", "EAFE P/E percentile. Threshold: < 13x cheap, > 19x expensive"),
    ("dm_equity",    "erp_us",         "V", "+", "30%", "DM earnings yield vs bond yield. Higher than US ERP = more attractive."),
    ("dm_equity",    "pe_spx",         "V", "+", "25%", "EAFE P/E / S&P P/E z-score. Low = DM cheap vs US = bullish DM."),
    ("dm_equity",    "shiller_cape",   "V", "+", "10%", "DM div yield vs US div yield (carry advantage)"),

    # =========================================================================
    # EM EQUITY (broad)
    # =========================================================================
    ("em_equity",    "pmi_china_mfg",  "F", "+", "30%", "China = 30%+ of MSCI EM. China PMI is dominant EM fundamental."),
    ("em_equity",    "cesi_em",        "F", "+", "25%", "Broad EM economic surprises. Moderate range (not extreme contrarian)."),
    ("em_equity",    "gdp_em_rev",     "F", "+", "20%", "EM growth consensus change. Upward revision = bullish EM earnings"),
    ("em_equity",    "eps_em_fwd",     "F", "+", "15%", "EM earnings growth trend. > 10% = positive signal"),
    ("em_equity",    "eps_beat_us",    "F", "+", "10%", "EM delivery vs estimates"),
    ("em_equity",    "em_mom",         "M", "+", "35%", "MSCI EM price momentum. Primary signal."),
    ("em_equity",    "em_xcn_mom",     "M", "+", "20%", "EM ex-China breadth check. Both confirming = high conviction."),
    ("em_equity",    "china_mom",      "M", "+", "15%", "China equity momentum - often diverges from broad EM"),
    ("em_equity",    "em_oas_mom",     "M", "+", "20%", "EM credit conditions leading EM equity by 2-4 weeks"),
    ("em_equity",    "dxy_mom",        "M", "+", "10%", "Dollar weakness = EM equity bull case."),
    ("em_equity",    "dxy_level",      "S", "-", "30%", "The #1 headwind for EM equity. z_DXY sign inverted."),
    ("em_equity",    "vix_level",      "S", "C", "25%", "High VIX = risk-off = EM equity suffers most (high beta to global risk)"),
    ("em_equity",    "embi_spread",    "S", "-", "25%", "EM sovereign stress is directly negative for EM equity"),
    ("em_equity",    "aaii_bb",        "S", "C", "20%", "Institutional positioning proxy (via retail sentiment)"),
    ("em_equity",    "pe_em",          "V", "+", "30%", "NDUEEGF PE percentile. < 10x very cheap, > 14x expensive."),
    ("em_equity",    "pe_spx",         "V", "+", "25%", "EM P/E / S&P P/E z-score. Historically EM trades at 25-35% discount."),
    ("em_equity",    "erp_em",         "V", "+", "25%", "EM risk-adjusted ERP. Accounts for sovereign risk premium."),
    ("em_equity",    "em_oas_lv",      "V", "+", "20%", "EM credit cheap = better financing conditions = positive for EM equity"),

    # =========================================================================
    # EM ex-CHINA
    # =========================================================================
    ("em_xchina",    "pmi_china_mfg",  "F", "+", "25%", "Regional spillovers still matter, but weaker for ex-China"),
    ("em_xchina",    "cesi_em",        "F", "+", "25%", "Broad EM surprises"),
    ("em_xchina",    "gdp_em_rev",     "F", "+", "25%", "EM consensus growth revisions"),
    ("em_xchina",    "eps_em_fwd",     "F", "+", "15%", "EM ex-China earnings trend"),
    ("em_xchina",    "eps_beat_us",    "F", "+", "10%", "EM delivery proxy"),
    ("em_xchina",    "em_xcn_mom",     "M", "+", "50%", "MSCI EM ex-China TR composite"),
    ("em_xchina",    "em_mom",         "M", "+", "20%", "Broad EM context"),
    ("em_xchina",    "em_oas_mom",     "M", "+", "20%", "EM credit leads equity"),
    ("em_xchina",    "dxy_mom",        "M", "+", "10%", "USD weakness supportive"),
    ("em_xchina",    "dxy_level",      "S", "-", "35%", "Strong USD hurts EM ex-China"),
    ("em_xchina",    "vix_level",      "S", "C", "25%", "Global risk regime"),
    ("em_xchina",    "embi_spread",    "S", "-", "25%", "EM sovereign stress"),
    ("em_xchina",    "aaii_bb",        "S", "C", "15%", "Sentiment proxy"),
    ("em_xchina",    "pe_em_xcn",      "V", "+", "40%", "EM ex-China P/E percentile"),
    ("em_xchina",    "erp_em",         "V", "+", "30%", "Risk-adjusted ERP"),
    ("em_xchina",    "em_oas_lv",      "V", "+", "30%", "EM credit valuation"),

    # =========================================================================
    # CHINA EQUITY
    # =========================================================================
    ("china_equity", "pmi_china_mfg",  "F", "++","35%", "Primary China activity signal. Level > 50 + direction rising = maximum bullish."),
    ("china_equity", "pmi_china_svcs", "F", "+", "20%", "China consumer and services economy. Services > 50% of GDP."),
    ("china_equity", "gdp_china_rev",  "F", "+", "20%", "Bloomberg consensus China GDP. Revision direction is key."),
    ("china_equity", "cesi_china",     "F", "+", "15%", "China economic data vs expectations"),
    ("china_equity", "eps_china_fwd",  "F", "+", "10%", "MSCI China earnings growth trend and revision"),
    ("china_equity", "china_mom",      "M", "++","40%", "China equity price momentum. Note: can diverge sharply from EM broad."),
    ("china_equity", "spxt_ma",        "M", "+", "25%", "China trend signal (MA proxy)"),
    ("china_equity", "em_xcn_mom",     "M", "+", "20%", "Is China outperforming or underperforming rest of EM?"),
    ("china_equity", "em_oas_mom",     "M", "+", "15%", "EM credit as financial conditions proxy for China equity"),
    ("china_equity", "dxy_level",      "S", "-", "30%", "CNY weakness proxy via DXY"),
    ("china_equity", "embi_spread",    "S", "-", "30%", "China CDS or EMBI component for sovereign risk"),
    ("china_equity", "vix_level",      "S", "C", "40%", "China is highest beta within EM to global risk-off events"),
    ("china_equity", "pe_china",       "V", "+", "35%", "NDEUCHF PE percentile. China often < 10x = cheap. < 8x = very cheap."),
    ("china_equity", "pe_spx",         "V", "+", "30%", "China P/E / S&P P/E. Historically 35-55% discount."),
    ("china_equity", "erp_em",         "V", "+", "25%", "Risk-adjusted ERP accounting for geopolitical/political risk"),
    ("china_equity", "em_oas_lv",      "V", "+", "10%", "Financing conditions for Chinese corporates"),
]

def _sheet_signal_mapping(wb):
    ws = wb.create_sheet("SignalMapping")
    ws.sheet_properties.tabColor = "22C55E"
    headers = ["ac_id","series_id","pillar","sign","weight_in_pillar","description_override"]
    ws.append(headers)
    for row in SIGNAL_MAPPING:
        ws.append(list(row))
    _apply_header(ws)
    _autosize(ws, max_w=80)
    ws.freeze_panes = "A2"
    return ws

# =============================================================================
# MAIN
# =============================================================================
def main():
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)
    _sheet_instructions(wb)
    _sheet_asset_classes(wb)
    _sheet_data_series(wb)
    _sheet_pillar_weights(wb)
    _sheet_pillar_notes(wb)
    _sheet_signal_mapping(wb)
    wb.save(OUT)
    print(f"OK  wrote {OUT}")
    print(f"    AssetClasses:  {len(ASSET_CLASSES)} rows")
    print(f"    DataSeries:    {len(DATA_SERIES)} rows")
    print(f"    PillarWeights: {len(PILLAR_WEIGHTS)} rows")
    print(f"    PillarNotes:   {len(PILLAR_NOTES)} rows")
    print(f"    SignalMapping: {len(SIGNAL_MAPPING)} rows")

if __name__ == "__main__":
    main()
