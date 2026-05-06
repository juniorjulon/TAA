"""
src/rebuild_signal_mapping.py
==============================
One-time script to:
  1. Remove em_xchina and china_equity from AssetClasses + PillarWeights sheets
  2. Add 5 new DataSeries rows (H7 series: breakeven_1y, gdpnow, nfci, fci_ez, fci_uk)
  3. Completely replace SignalMapping with a clean, duplicate-free, no-orphan set

Run once:
  python src/rebuild_signal_mapping.py

Plain-vanilla design philosophy:
  - Every signal has real data (verified against executable DataSeries list)
  - No duplicate series within same (AC, pillar)
  - Sign and economic logic are obvious and explainable to a committee
  - Complex/uncertain relationships documented in docs/signal_improvements.md
"""

import os, sys, warnings
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
warnings.filterwarnings("ignore")

from openpyxl import load_workbook
from config import CONFIG_XLSX

XLSX = CONFIG_XLSX

# ─────────────────────────────────────────────────────────────────────────────
# 5 new DataSeries rows (H7 sheet)
# ─────────────────────────────────────────────────────────────────────────────

NEW_DS = [
    # (series_id, signal_name, ticker, source, frequency, pillar,
    #  transformation, window, notes, series_type, input_sheet, input_column, transform_code)
    ("breakeven_1y", "US 1Y Breakeven Inflation",
     "USGGBE01 Index", "Bloomberg", "daily", "Fundamentals",
     "Rolling z-score 3Y", "756",
     "1Y inflation expectation — most relevant for short-end FI fundamentals",
     "original", "H7", "USGGBE01 Index", "rolling_z"),

    ("gdpnow", "Atlanta Fed GDPNow US Forecast",
     "GDGCAFJP Index", "Bloomberg/FRB Atlanta", "daily", "Fundamentals",
     "EWMA z-score", "252",
     "Real-time US GDP tracking estimate; high frequency complement to consensus GDP_US",
     "original", "H7", "GDGCAFJP Index", "ewma_z"),

    ("nfci", "Chicago Fed National Financial Conditions Index",
     "NFCIINDX Index", "Bloomberg/FRB Chicago", "daily", "Sentiment",
     "EWMA z-score", "252*3",
     "Broad US financial conditions; negative = tight = headwind for risk assets",
     "original", "H7", "NFCIINDX Index", "ewma_z"),

    ("fci_ez", "Bloomberg Euro-Zone Financial Conditions Index",
     "BFCIEU Index", "Bloomberg", "daily", "Sentiment",
     "EWMA z-score", "252*3",
     "EZ financial conditions; tight = headwind for DM equity (Europe-focused)",
     "original", "H7", "BFCIEU Index", "ewma_z"),

    ("fci_uk", "Bloomberg UK Financial Conditions Index",
     "BFCIGB Index", "Bloomberg", "daily", "Sentiment",
     "EWMA z-score", "252*3",
     "UK financial conditions; available for future DM decomposition",
     "original", "H7", "BFCIGB Index", "ewma_z"),
]

# ─────────────────────────────────────────────────────────────────────────────
# COMPLETE SignalMapping — plain vanilla, all series executable, no duplicates
# Format: (ac_id, series_id, pillar, sign, weight, description)
#
# Weight convention: decimals (e.g. 0.30 = 30%); _wavg() renormalizes so they
# don't need to sum to exactly 1.0 per (AC, pillar).
#
# Sign convention:
#   +1 = series positive → signal bullish for AC
#   -1 = series positive → signal bearish for AC
# ─────────────────────────────────────────────────────────────────────────────

def _r(ac, sid, pillar, sign, w, desc=""):
    return (ac, sid, pillar, sign, w, desc)

SM = [
    # ══════════════════════════════════════════════════════════════════════════
    # MONEY MARKET
    # ══════════════════════════════════════════════════════════════════════════
    # F: Rate environment — high real rates = attractive MM yield
    _r("money_market", "real_ff",       "F", +1, 0.50, "Real Fed Funds = FDTR - PCE; positive + high = MM yield attractive"),
    _r("money_market", "core_pce",      "F", +1, 0.50, "PCE above 2% = Fed holds rates higher = positive for MM rates"),

    # M: Short-end price momentum
    _r("money_market", "lt03_price",    "M", +1, 0.60, "Short-term govt TR momentum"),
    _r("money_market", "gt02_mom",      "M", +1, 0.40, "2Y yield rising = positive momentum for cash return"),

    # S: Safe-haven / risk-off demand for cash
    _r("money_market", "modern_ted",    "S", +1, 0.40, "Funding stress (tbill-SOFR spread) = demand for cash"),
    _r("money_market", "hy_safe_haven", "S", +1, 0.35, "HY stress proxy = flight to cash/safety"),
    _r("money_market", "vix",           "S", +1, 0.25, "High VIX = risk-off = prefer cash equivalents"),

    # V: Yield carry attractiveness
    _r("money_market", "gt02",          "V", +1, 0.50, "2Y yield pctile: high = attractive carry for MM"),
    _r("money_market", "tips_5y",       "V", +1, 0.25, "TIPS 5Y real yield: positive = real return attractive"),
    _r("money_market", "term_spread",   "V", -1, 0.25, "Flat/inverted curve = short-end carry attractive (inverted sign)"),

    # ══════════════════════════════════════════════════════════════════════════
    # SHORT-TERM FIXED INCOME (dur ~0.5Y, similar to MM but with credit carry)
    # ══════════════════════════════════════════════════════════════════════════
    # F: Growth/inflation bearish for short duration (growth = higher s-t rates)
    _r("short_term_fi", "pmi_us",       "F", -1, 0.30, "PMI up = rates rise = headwind for ST duration"),
    _r("short_term_fi", "gdp_us",       "F", -1, 0.25, "GDP growth = rates stay higher for longer"),
    _r("short_term_fi", "core_pce",     "F", -1, 0.20, "Inflation above target = Fed restricts = ST rate headwind"),
    _r("short_term_fi", "breakeven_1y", "F", -1, 0.25, "1Y inflation expectation most relevant for short-end"),

    # M: Price momentum — short-term bond indices
    _r("short_term_fi", "bfu5_price",   "M", +1, 0.40, "Short-term IG corp TR momentum (primary)"),
    _r("short_term_fi", "i132_price",   "M", +1, 0.25, "Alternative short-term corp TR momentum"),
    _r("short_term_fi", "gt02_mom",     "M", +1, 0.20, "2Y yield direction: falling = positive for duration"),
    _r("short_term_fi", "oas_bbb_mom",  "M", +1, 0.15, "IG spread tightening = positive for credit return"),

    # S: Safe-haven / risk regime
    _r("short_term_fi", "modern_ted",   "S", +1, 0.40, "Funding stress = flight to quality ST FI"),
    _r("short_term_fi", "move_z",       "S", +1, 0.35, "Bond vol rising = safe-haven demand for duration"),
    _r("short_term_fi", "vix",          "S", +1, 0.25, "Risk-off = prefer short-duration FI"),

    # V: Yield and spread carry attractiveness
    _r("short_term_fi", "gt02",         "V", +1, 0.40, "2Y yield pctile: high = ST FI attractive"),
    _r("short_term_fi", "oas_bbb",      "V", +1, 0.30, "BBB OAS pctile: wide = cheap IG carry"),
    _r("short_term_fi", "tips_5y",      "V", +1, 0.30, "TIPS 5Y real yield"),

    # ══════════════════════════════════════════════════════════════════════════
    # LT US TREASURIES
    # ══════════════════════════════════════════════════════════════════════════
    # F: Macro bearish for duration (growth/inflation = higher rates)
    _r("lt_treasuries", "pmi_us",       "F", -1, 0.30, "PMI up = rates rise = UST prices fall"),
    _r("lt_treasuries", "gdp_us",       "F", -1, 0.25, "GDP growth = Fed stays restrictive = duration headwind"),
    _r("lt_treasuries", "cesi_us",      "F", -1, 0.20, "Better-than-expected data = rates rise"),
    _r("lt_treasuries", "breakeven_10y","F", -1, 0.25, "10Y inflation expectations: rising = UST real yield falls"),

    # M: Price and yield momentum
    _r("lt_treasuries", "bsgv_price",   "M", +1, 0.45, "LT Govt TR momentum — primary"),
    _r("lt_treasuries", "gt10_mom",     "M", +1, 0.35, "10Y yield falling = positive for UST"),
    _r("lt_treasuries", "oas_bbb_mom",  "M", +1, 0.20, "Credit stress = flight-to-quality = UST demand"),

    # S: Safe-haven demand (all three point same direction for UST)
    _r("lt_treasuries", "vix",          "S", +1, 0.40, "High VIX = risk-off = UST safe-haven demand"),
    _r("lt_treasuries", "move_z",       "S", +1, 0.30, "MOVE rising = UST vol = safe-haven hedging demand"),
    _r("lt_treasuries", "modern_ted",   "S", +1, 0.20, "Funding stress = flight to UST"),
    _r("lt_treasuries", "skew_z",       "S", +1, 0.10, "High SKEW = tail-risk hedging = UST safe-haven"),

    # V: Yield carry attractiveness
    _r("lt_treasuries", "gt10",         "V", +1, 0.40, "10Y yield pctile: high = LT UST attractive carry"),
    _r("lt_treasuries", "tips_10y",     "V", +1, 0.35, "TIPS 10Y real yield: positive and high = UST very attractive"),
    _r("lt_treasuries", "term_spread",  "V", +1, 0.25, "Steep curve = long-end premium = LT attractive"),

    # ══════════════════════════════════════════════════════════════════════════
    # LT US CORPORATE (IG)
    # ══════════════════════════════════════════════════════════════════════════
    # F: Growth/earnings bullish for credit (growth = tighter spreads)
    _r("lt_us_corp", "pmi_us",          "F", +1, 0.25, "ISM PMI up = corporate health = spreads tighten"),
    _r("lt_us_corp", "cesi_us",         "F", +1, 0.20, "Positive surprises = better credit conditions"),
    _r("lt_us_corp", "gdp_us",          "F", +1, 0.20, "GDP growth = earnings stable = credit positive"),
    _r("lt_us_corp", "eps_us",          "F", +1, 0.25, "EPS revisions = direct corporate earnings signal"),
    _r("lt_us_corp", "eps_rev_us",      "F", +1, 0.10, "Multi-horizon EPS revision (3M+6M weighted)"),

    # M: Spread and price momentum
    _r("lt_us_corp", "bfu5_price",      "M", +1, 0.25, "IG corp TR price momentum"),
    _r("lt_us_corp", "oas_bbb_mom",     "M", +1, 0.40, "BBB spread tightening = credit risk appetite"),
    _r("lt_us_corp", "oas_hy_mom",      "M", +1, 0.20, "HY spread tightening = broader credit cycle signal"),
    _r("lt_us_corp", "gt10_mom",        "M", +1, 0.15, "Duration component: falling 10Y = LT corp price up"),

    # S: Credit risk regime (high stress = negative for IG corp)
    _r("lt_us_corp", "vix",             "S", -1, 0.35, "Risk-on = positive for credit; high VIX = spread widening"),
    _r("lt_us_corp", "hy_stress",       "S", -1, 0.35, "HY widening proxy = credit headwind"),
    _r("lt_us_corp", "modern_ted",      "S", -1, 0.30, "Funding stress = credit spreads widen = negative for IG"),

    # V: Spread valuation (carry attractiveness)
    _r("lt_us_corp", "oas_bbb",         "V", +1, 0.40, "BBB OAS pctile: wide = attractive IG carry"),
    _r("lt_us_corp", "oas_hy",          "V", +1, 0.25, "HY OAS pctile: broad credit valuation context"),
    _r("lt_us_corp", "hy_ig_ratio",     "V", +1, 0.15, "HY/IG ratio: high = HY cheap vs IG = rotate to IG"),
    _r("lt_us_corp", "gt10",            "V", +1, 0.20, "10Y yield level: duration carry component"),

    # ══════════════════════════════════════════════════════════════════════════
    # LT EM FIXED INCOME (CEMBI IG BBB — user's main LT FI holding)
    # ══════════════════════════════════════════════════════════════════════════
    # F: EM macro fundamentals (bullish = spreads tighten = EM credit positive)
    _r("lt_em_fi", "pmi_china",         "F", +1, 0.30, "China PMI leads EM credit broadly (China = ~25% CEMBI)"),
    _r("lt_em_fi", "gdp_em",            "F", +1, 0.30, "EM growth consensus: stronger = EM credit positive"),
    _r("lt_em_fi", "cesi_em",           "F", +1, 0.20, "EM economic surprises: positive = tighter spreads"),
    _r("lt_em_fi", "eps_em",            "F", +1, 0.20, "EM EPS revisions: earnings health = credit health"),

    # M: EM spread and equity momentum
    _r("lt_em_fi", "oas_em_mom",        "M", +1, 0.45, "EM spread tightening = primary EM credit momentum"),
    _r("lt_em_fi", "msci_em_tr",        "M", +1, 0.35, "EM equity momentum: risk appetite proxy for EM credit"),
    _r("lt_em_fi", "oas_bbb_mom",       "M", +1, 0.20, "Global IG momentum: contagion from DM credit cycle"),

    # S: EM risk regime (stress signals negative for EM credit)
    _r("lt_em_fi", "embi",              "S", -1, 0.40, "EM sovereign stress proxy (OAS_EM.diff(21)); widening = risk-off"),
    _r("lt_em_fi", "em_stress",         "S", -1, 0.35, "EM-specific OAS stress: widening = EM credit headwind"),
    _r("lt_em_fi", "vix",               "S", -1, 0.25, "Global risk-off = EM spread widening = negative for EM credit"),

    # V: EM spread carry attractiveness
    _r("lt_em_fi", "oas_em",            "V", +1, 0.50, "EM BBB OAS pctile: wide = attractive carry (primary)"),
    _r("lt_em_fi", "oas_latam",         "V", +1, 0.25, "LatAm OAS pctile: regional carry component"),
    _r("lt_em_fi", "gt10",              "V", +1, 0.25, "10Y UST yield: benchmark rate — high = EM credit attractive vs UST"),

    # ══════════════════════════════════════════════════════════════════════════
    # US EQUITY (BROAD)
    # ══════════════════════════════════════════════════════════════════════════
    # F: US macro and earnings (growth = bullish for equity)
    _r("us_equity", "pmi_us",           "F", +1, 0.25, "ISM PMI: leading business cycle indicator"),
    _r("us_equity", "cesi_us",          "F", +1, 0.20, "Economic surprises: beats = positive earnings revisions"),
    _r("us_equity", "gdp_us",           "F", +1, 0.20, "GDP consensus growth"),
    _r("us_equity", "gdpnow",           "F", +1, 0.10, "Atlanta Fed real-time GDP tracking: high-frequency complement"),
    _r("us_equity", "eps_us",           "F", +1, 0.15, "EPS revisions (1M)"),
    _r("us_equity", "eps_rev_us",       "F", +1, 0.10, "EPS revisions (3M+6M): stronger multi-horizon signal"),

    # M: Price and credit risk appetite momentum
    _r("us_equity", "sp500_tr",         "M", +1, 0.60, "S&P 500 TR price momentum: primary equity momentum"),
    _r("us_equity", "oas_hy_mom",       "M", +1, 0.40, "HY spread tightening = credit/risk appetite = equity positive"),

    # S: Sentiment (contrarian at extremes)
    _r("us_equity", "vix",              "S", +1, 0.40, "VIX contrarian: high fear = buy signal for equity"),
    _r("us_equity", "aaii_z",           "S", -1, 0.35, "AAII bull-bear: extreme bullish = crowded = sell signal"),
    _r("us_equity", "fci_z",            "S", -1, 0.25, "Bloomberg US FCI: tight conditions = equity headwind"),

    # V: Equity valuation
    _r("us_equity", "pe_score_sp500",   "V", +1, 0.40, "S&P 500 PE percentile: cheap = positive"),
    _r("us_equity", "erp_us",           "V", +1, 0.40, "US ERP (S&P EY - TIPS 10Y): most powerful cross-asset signal"),
    _r("us_equity", "rel_pe_us_em",     "V", +1, 0.20, "Relative PE: log(PE_EM / PE_US) — US cheap vs EM = positive"),

    # ══════════════════════════════════════════════════════════════════════════
    # US GROWTH EQUITY
    # ══════════════════════════════════════════════════════════════════════════
    # F: Same macro backdrop as US Equity
    _r("us_growth", "pmi_us",           "F", +1, 0.25, "Business cycle drives growth premium"),
    _r("us_growth", "cesi_us",          "F", +1, 0.20, "Positive surprises amplify growth EPS revisions"),
    _r("us_growth", "gdp_us",           "F", +1, 0.20, "GDP growth supports premium valuations"),
    _r("us_growth", "gdpnow",           "F", +1, 0.10, "Real-time GDP tracking"),
    _r("us_growth", "eps_us",           "F", +1, 0.15, "EPS revision momentum"),
    _r("us_growth", "eps_rev_us",       "F", +1, 0.10, "Multi-horizon EPS revision"),

    # M: Growth index price momentum only (plain vanilla)
    _r("us_growth", "sp500_gro_tr",     "M", +1, 1.00, "S&P 500 Growth TR: primary momentum signal"),

    # S: Same sentiment as US Equity
    _r("us_growth", "vix",              "S", +1, 0.40, "VIX contrarian"),
    _r("us_growth", "aaii_z",           "S", -1, 0.35, "AAII contrarian: extreme bullish = sell"),
    _r("us_growth", "fci_z",            "S", -1, 0.25, "Tight FCI = growth multiple compression headwind"),

    # V: Growth-specific valuation
    _r("us_growth", "pe_score_gro",     "V", +1, 0.40, "Growth PE percentile: cheap growth = positive"),
    _r("us_growth", "erp_us",           "V", +1, 0.35, "US ERP: applies equally to growth style"),
    _r("us_growth", "rel_pe_gro_val",   "V", +1, 0.25, "Growth vs Value relative PE: log(PE_Value/PE_Growth)"),

    # ══════════════════════════════════════════════════════════════════════════
    # US VALUE EQUITY
    # ══════════════════════════════════════════════════════════════════════════
    # F: Same macro as US Equity
    _r("us_value", "pmi_us",            "F", +1, 0.25, "Business cycle: value benefits from cyclical recovery"),
    _r("us_value", "cesi_us",           "F", +1, 0.20, "Positive surprises benefit value (often cyclical sectors)"),
    _r("us_value", "gdp_us",            "F", +1, 0.20, "GDP growth supports value (financials, industrials, energy)"),
    _r("us_value", "gdpnow",            "F", +1, 0.10, "Real-time GDP tracking"),
    _r("us_value", "eps_us",            "F", +1, 0.15, "EPS revision momentum"),
    _r("us_value", "eps_rev_us",        "F", +1, 0.10, "Multi-horizon EPS revision"),

    # M: Value index price momentum only (plain vanilla)
    _r("us_value", "sp500_val_tr",      "M", +1, 1.00, "S&P 500 Value TR: primary momentum signal"),

    # S: Same sentiment as US Equity
    _r("us_value", "vix",               "S", +1, 0.40, "VIX contrarian"),
    _r("us_value", "aaii_z",            "S", -1, 0.35, "AAII contrarian: extreme bullish = sell"),
    _r("us_value", "fci_z",             "S", -1, 0.25, "Tight FCI = credit conditions headwind (value is credit-sensitive)"),

    # V: Value-specific valuation
    _r("us_value", "pe_score_val",      "V", +1, 0.40, "Value PE percentile: cheap value = positive"),
    _r("us_value", "erp_us",            "V", +1, 0.35, "US ERP: ERP expansion = value re-rates first"),
    _r("us_value", "rel_pe_val_gro",    "V", +1, 0.25, "Value vs Growth relative PE: log(PE_Growth/PE_Value)"),

    # ══════════════════════════════════════════════════════════════════════════
    # DM EX-US EQUITY
    # ══════════════════════════════════════════════════════════════════════════
    # F: European and DM macro fundamentals
    _r("dm_equity", "pmi_ez",           "F", +1, 0.30, "EZ PMI: primary DM activity indicator (Europe ~60% EAFE)"),
    _r("dm_equity", "cesi_ez",          "F", +1, 0.25, "EZ economic surprises"),
    _r("dm_equity", "gdp_dm",           "F", +1, 0.25, "DM GDP consensus: broad developed market growth"),
    _r("dm_equity", "eps_eafe",         "F", +1, 0.20, "EAFE EPS revisions: DM corporate earnings signal"),

    # M: EAFE price momentum (plain vanilla: one clean index)
    _r("dm_equity", "eafe_tr",          "M", +1, 1.00, "MSCI EAFE TR: primary DM equity momentum signal"),

    # S: DM sentiment
    _r("dm_equity", "vix",              "S", +1, 0.35, "VIX contrarian: global risk-off = DM beaten down"),
    _r("dm_equity", "vstoxx_z",         "S", +1, 0.35, "VSTOXX: European vol contrarian (DM-specific fear gauge)"),
    _r("dm_equity", "fci_ez",           "S", -1, 0.30, "EZ financial conditions: tight = DM equity headwind"),

    # V: DM equity valuation
    _r("dm_equity", "pe_score_eafe",    "V", +1, 0.40, "EAFE PE percentile: cheap DM = positive"),
    _r("dm_equity", "erp_acwi",         "V", +1, 0.35, "ACWI ERP (EY - TIPS 10Y): cross-asset valuation"),
    _r("dm_equity", "rel_pe_dm_us",     "V", +1, 0.25, "DM vs US relative PE: log(PE_US/PE_EAFE) — DM cheap vs US = positive"),

    # ══════════════════════════════════════════════════════════════════════════
    # EMERGING MARKETS EQUITY
    # ══════════════════════════════════════════════════════════════════════════
    # F: EM macro fundamentals
    _r("em_equity", "pmi_china",        "F", +1, 0.30, "China PMI: most influential EM growth indicator"),
    _r("em_equity", "gdp_em",           "F", +1, 0.25, "EM GDP consensus growth"),
    _r("em_equity", "cesi_em",          "F", +1, 0.25, "EM economic surprises"),
    _r("em_equity", "eps_em",           "F", +1, 0.20, "EM EPS revisions: corporate health in EM"),

    # M: EM price and spread momentum
    _r("em_equity", "msci_em_tr",       "M", +1, 0.55, "MSCI EM TR: primary EM equity momentum"),
    _r("em_equity", "oas_em_mom",       "M", +1, 0.45, "EM spread tightening = risk appetite = EM equity positive"),

    # S: EM sentiment / stress
    _r("em_equity", "embi",             "S", -1, 0.40, "EM sovereign stress proxy: widening = EM risk-off"),
    _r("em_equity", "vix",              "S", +1, 0.30, "Global VIX contrarian: high fear = EM beaten down"),
    _r("em_equity", "em_stress",        "S", -1, 0.30, "EM-specific OAS stress proxy: widening = headwind"),

    # V: EM equity valuation
    _r("em_equity", "pe_score_em",      "V", +1, 0.35, "MSCI EM PE percentile: cheap EM = positive"),
    _r("em_equity", "erp_em",           "V", +1, 0.40, "EM ERP (EM EY - TIPS 10Y): EM risk premium attractiveness"),
    _r("em_equity", "rel_pe_em_us",     "V", +1, 0.25, "EM vs US relative PE: log(PE_US/PE_EM) — EM cheap vs US"),
]


def main():
    wb = load_workbook(XLSX)

    # ── 1. Remove em_xchina and china_equity from AssetClasses sheet ──────────
    REMOVE_ACS = {"em_xchina", "china_equity"}

    for sheet_name in ["AssetClasses", "PillarWeights"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows_to_delete = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row and row[0] and str(row[0]).strip() in REMOVE_ACS:
                rows_to_delete.append(i)
        for i in reversed(rows_to_delete):
            ws.delete_rows(i)
        print(f"  {sheet_name}: removed {len(rows_to_delete)} rows ({REMOVE_ACS})")

    # ── 2. Add 5 new DataSeries rows (skip if already exist) ─────────────────
    ws_ds = wb["DataSeries"]
    ds_rows = list(ws_ds.values)
    ds_headers = [str(h).strip() if h else f"c{i}" for i, h in enumerate(ds_rows[0])]
    existing_ids = {str(r[0]).strip() for r in ds_rows[1:] if r and r[0]}

    idx_type  = ds_headers.index("series_type")   if "series_type"   in ds_headers else None
    idx_sheet = ds_headers.index("input_sheet")   if "input_sheet"   in ds_headers else None
    idx_col   = ds_headers.index("input_column")  if "input_column"  in ds_headers else None
    idx_trans = ds_headers.index("transform_code") if "transform_code" in ds_headers else None

    added_ds = 0
    for row in NEW_DS:
        sid = row[0]
        if sid in existing_ids:
            print(f"  DataSeries: SKIP (exists) {sid}")
            continue
        base = list(row[:9]) + [""] * (len(ds_headers) - 9)
        if idx_type  is not None: base[idx_type]  = row[9]
        if idx_sheet is not None: base[idx_sheet] = row[10]
        if idx_col   is not None: base[idx_col]   = row[11]
        if idx_trans is not None: base[idx_trans] = row[12]
        ws_ds.append(base)
        added_ds += 1
        print(f"  DataSeries: ADD {sid}")

    # ── 3. Rebuild SignalMapping from scratch ─────────────────────────────────
    ws_sm = wb["SignalMapping"]
    # Delete all data rows (keep header row 1)
    max_row = ws_sm.max_row
    if max_row > 1:
        ws_sm.delete_rows(2, max_row - 1)
    print(f"  SignalMapping: cleared {max_row - 1} old rows")

    # Write fresh rows
    for (ac, sid, pillar, sign, weight, desc) in SM:
        ws_sm.append([ac, sid, pillar, sign, weight, desc])
    print(f"  SignalMapping: wrote {len(SM)} new rows")

    wb.save(XLSX)
    print(f"\nDone: {added_ds} new DataSeries, {len(SM)} SignalMapping rows -> {XLSX}")

    # Summary: signals per AC
    from collections import Counter
    ac_counts = Counter(r[0] for r in SM)
    pillar_counts = Counter((r[0], r[2]) for r in SM)
    print("\nSignals per AC:")
    for ac, cnt in sorted(ac_counts.items()):
        print(f"  {ac:<18}: {cnt} total | " +
              " | ".join(f"{p}={pillar_counts[(ac,p)]}"
                         for p in "FMSV"))


if __name__ == "__main__":
    main()
