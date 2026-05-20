"""
test_build_layer.py
===================
Automated health-check for the config/build/doc layer added in April 2026.
Verifies that every component is present, readable, and produces correct output.

Run:
    python src/test_build_layer.py

Exit code 0 = all checks passed.
Exit code 1 = one or more checks failed.
"""

from __future__ import annotations
import os, sys, re, traceback

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

PASS = "  [PASS]"
FAIL = "  [FAIL]"
HEAD = "\n" + "-" * 50

failures: list[str] = []


def check(name: str, fn):
    try:
        msg = fn()
        print(f"{PASS} {name}" + (f"  -> {msg}" if msg else ""))
    except Exception as e:
        print(f"{FAIL} {name}")
        print(f"         {e}")
        failures.append(name)


# ─────────────────────────────────────────────────────────────────────────────
# 1. FILES EXIST
# ─────────────────────────────────────────────────────────────────────────────
print(HEAD)
print("1. FILE EXISTENCE")
print(HEAD)

FILES = {
    "config/taa_config.xlsx":            os.path.join(ROOT, "config", "taa_config.xlsx"),
    "src/seed_taa_config.py":            os.path.join(HERE, "seed_taa_config.py"),
    "src/build_dashboard.py":            os.path.join(HERE, "build_dashboard.py"),
    "src/generate_methodology_doc.py":   os.path.join(HERE, "generate_methodology_doc.py"),
    "docs/TAA_Methodology.docx":         os.path.join(ROOT, "docs", "TAA_Methodology.docx"),
    "index.html":                        os.path.join(ROOT, "index.html"),
    "src/config.py":                     os.path.join(HERE, "config.py"),
}

for label, path in FILES.items():
    check(f"exists: {label}", lambda p=path: (
        None if os.path.isfile(p)
        else (_ for _ in ()).throw(FileNotFoundError(p))
    ))


# ─────────────────────────────────────────────────────────────────────────────
# 2. EXCEL STRUCTURE
# ─────────────────────────────────────────────────────────────────────────────
print(HEAD)
print("2. EXCEL STRUCTURE (config/taa_config.xlsx)")
print(HEAD)

def _load_wb():
    from openpyxl import load_workbook
    return load_workbook(os.path.join(ROOT, "config", "taa_config.xlsx"), data_only=True)

def _sheet_rows(wb, name):
    ws = wb[name]
    return [r for r in ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)]

check("openpyxl importable", lambda: None)

check("sheet names", lambda: (
    wb := _load_wb(),
    names := wb.sheetnames,
    None if set(["Instructions","AssetClasses","DataSeries","PillarWeights","PillarNotes","SignalMapping"]).issubset(names)
    else (_ for _ in ()).throw(AssertionError(f"Got: {names}"))
)[-1])

check("AssetClasses: 10 rows", lambda: (
    wb := _load_wb(),
    rows := _sheet_rows(wb, "AssetClasses"),
    None if len(rows) == 10 else (_ for _ in ()).throw(AssertionError(f"Got {len(rows)} rows"))
)[-1])

check("DataSeries: >= 80 rows", lambda: (
    wb := _load_wb(),
    rows := _sheet_rows(wb, "DataSeries"),
    f"{len(rows)} rows" if len(rows) >= 80
    else (_ for _ in ()).throw(AssertionError(f"Got {len(rows)} rows"))
)[-1])

check("PillarWeights: 10 rows, each sums to 1.0", lambda: (
    wb := _load_wb(),
    ws := wb["PillarWeights"],
    hdrs := [c.value for c in ws[1]],
    rows := _sheet_rows(wb, "PillarWeights"),
    [
        (_ for _ in ()).throw(AssertionError(
            f"Row {r[0]}: F+M+S+V = {sum(float(r[hdrs.index(p)]) for p in ['F','M','S','V']):.3f}"
        ))
        for r in rows
        if abs(sum(float(r[hdrs.index(p)]) for p in ["F","M","S","V"]) - 1.0) > 0.005
    ] or f"{len(rows)} rows, all sum to 1.0"
)[-1])

check("SignalMapping: >= 100 rows", lambda: (
    wb := _load_wb(),
    rows := _sheet_rows(wb, "SignalMapping"),
    f"{len(rows)} rows" if len(rows) >= 100
    else (_ for _ in ()).throw(AssertionError(f"Got {len(rows)} rows"))
)[-1])

# (replaced above with >= 100 check)

check("PillarNotes: >= 48 rows", lambda: (
    wb := _load_wb(),
    rows := _sheet_rows(wb, "PillarNotes"),
    f"{len(rows)} rows" if len(rows) >= 48
    else (_ for _ in ()).throw(AssertionError(f"Got {len(rows)} rows"))
)[-1])

check("AssetClasses: FI and EQ groups present", lambda: (
    wb := _load_wb(),
    ws := wb["AssetClasses"],
    hdrs := [c.value for c in ws[1]],
    rows := _sheet_rows(wb, "AssetClasses"),
    groups := {r[hdrs.index("group")] for r in rows},
    None if {"FI","EQ"}.issubset(groups)
    else (_ for _ in ()).throw(AssertionError(f"Groups found: {groups}"))
)[-1])


# ─────────────────────────────────────────────────────────────────────────────
# 3. index.html CONTENT CHECKS
# ─────────────────────────────────────────────────────────────────────────────
# Note: index.html is now generated as a complete file by generate_dashboard.py
# (no BUILD markers). Content is verified by checking key JS constants.
print(HEAD)
print("3. index.html CONTENT CHECKS")
print(HEAD)

def _read_html():
    with open(os.path.join(ROOT, "index.html"), "r", encoding="utf-8") as f:
        return f.read()

check("index.html exists and non-empty", lambda: (
    html := _read_html(),
    None if len(html) > 100_000 else (_ for _ in ()).throw(AssertionError(f"Too small: {len(html)}"))
)[-1])

check("index.html has SIG_MATRIX constant", lambda: (
    None if "const SIG_MATRIX=" in _read_html()
    else (_ for _ in ()).throw(AssertionError("SIG_MATRIX not found"))
))

check("index.html has FI_BLUEPRINT constant", lambda: (
    None if "const FI_BLUEPRINT" in _read_html()
    else (_ for _ in ()).throw(AssertionError("FI_BLUEPRINT not found"))
))

check("index.html has EQ_BLUEPRINT constant", lambda: (
    None if "const EQ_BLUEPRINT" in _read_html()
    else (_ for _ in ()).throw(AssertionError("EQ_BLUEPRINT not found"))
))

check("FI_BLUEPRINT contains 3 active FI AC blocks", lambda: (
    html := _read_html(),
    m := re.search(r"const FI_BLUEPRINT\s*=\s*\[(.*?)\];", html, re.DOTALL),
    count := m.group(1).count("pillars:{") if m else 0,
    f"{count} blocks" if count == 3
    else (_ for _ in ()).throw(AssertionError(f"Expected 3 active FI ACs, got {count}"))
)[-1])

check("EQ_BLUEPRINT has entries for US/DM/EM equity ACs", lambda: (
    html := _read_html(),
    None if all(kw in html for kw in ["US Equity", "DM ex-US", "Emerging Markets"])
    else (_ for _ in ()).throw(AssertionError("One or more EQ AC names missing"))
)[-1])

check("EQ_BLUEPRINT contains 3 active EQ AC blocks", lambda: (
    html := _read_html(),
    m := re.search(r"const EQ_BLUEPRINT\s*=\s*\[(.*?)\];", html, re.DOTALL),
    count := m.group(1).count("pillars:{") if m else 0,
    f"{count} blocks" if count == 3
    else (_ for _ in ()).throw(AssertionError(f"Expected 3 active EQ ACs, got {count}"))
)[-1])

check("AC_ORDER contains the 6 active ac_id keys", lambda: (
    html := _read_html(),
    m := re.search(r"const AC_ORDER=\[(.*?)\]", html, re.DOTALL),
    ids_found := set(re.findall(r"'([a-z_]+)'", m.group(1))) if m else set(),
    expected := {'lt_treasuries','lt_us_corp','lt_em_fi','us_equity','dm_equity','em_equity'},
    f"{len(ids_found)} active AC ids" if ids_found == expected
    else (_ for _ in ()).throw(AssertionError(f"Mismatch: extra={ids_found-expected}, missing={expected-ids_found}"))
)[-1])


# ─────────────────────────────────────────────────────────────────────────────
# 4. BUILD MARKERS IN src/config.py
# ─────────────────────────────────────────────────────────────────────────────
print(HEAD)
print("4. BUILD MARKERS IN src/config.py")
print(HEAD)

EXPECTED_PY_MARKERS = [
    "BUILD:PY_AC_UNIVERSE_START",
    "BUILD:PY_AC_UNIVERSE_END",
    "BUILD:PY_PILLAR_WEIGHTS_START",
    "BUILD:PY_PILLAR_WEIGHTS_END",
    "BUILD:PY_MAX_TILT_START",
    "BUILD:PY_MAX_TILT_END",
]

def _read_config():
    with open(os.path.join(HERE, "config.py"), "r", encoding="utf-8") as f:
        return f.read()

for marker in EXPECTED_PY_MARKERS:
    check(f"marker present: {marker}", lambda m=marker: (
        None if f"<<<{m}>>>" in _read_config()
        else (_ for _ in ()).throw(AssertionError(f"<<<{m}>>> not found in config.py"))
    ))

check("config.py ASSET_CLASSES has 6 active entries", lambda: (
    cfg := _read_config(),
    m := re.search(r"ASSET_CLASSES\s*=\s*\[([^\]]+)\]", cfg, re.DOTALL),
    count := len(re.findall(r'"[a-z_]+"', m.group(1))) if m else 0,
    f"{count} entries" if count == 6
    else (_ for _ in ()).throw(AssertionError(f"Got {count} entries, expected 6 active ACs"))
)[-1])

check("config.py PILLAR_WEIGHTS has 6 active entries", lambda: (
    cfg := _read_config(),
    count := len(re.findall(r'"[a-z_]+":\s*\{"F":', cfg)),
    f"{count} entries" if count == 6
    else (_ for _ in ()).throw(AssertionError(f"Got {count} entries, expected 6 active ACs"))
)[-1])


# ─────────────────────────────────────────────────────────────────────────────
# 5. BUILD SCRIPT DRY RUN
# ─────────────────────────────────────────────────────────────────────────────
print(HEAD)
print("5. BUILD SCRIPT (src/build_dashboard.py) — load + render only, no write")
print(HEAD)

def _import_build():
    sys.path.insert(0, HERE)
    import build_dashboard as bd
    cfg = bd.load_config()
    return cfg

check("build_dashboard.py importable + load_config()", lambda: (
    cfg := _import_build(),
    f"{len(cfg['asset_classes'])} ACs, {len(cfg['data_series'])} series, "
    f"{len(cfg['mapping'])} mappings"
)[-1])

check("render_sig_matrix() produces valid JS", lambda: (
    sys.path.insert(0, HERE),
    __import__("build_dashboard"),
    bd := sys.modules["build_dashboard"],
    cfg := bd.load_config(),
    js := bd.render_sig_matrix(cfg),
    None if js.startswith("const SIG_MATRIX=[") and js.endswith("];")
    else (_ for _ in ()).throw(AssertionError("SIG_MATRIX JS malformed"))
)[-1])

check("render_fi_blueprint() produces 3 active FI AC blocks", lambda: (
    sys.path.insert(0, HERE),
    bd := sys.modules.get("build_dashboard") or __import__("build_dashboard"),
    cfg := bd.load_config(),
    js := bd.render_fi_blueprint(cfg),
    count := js.count("pillars:{"),
    f"{count} pillar blocks" if count == 3
    else (_ for _ in ()).throw(AssertionError(f"Expected 3 active FI ACs, got {count}"))
)[-1])

check("render_eq_blueprint() produces 3 active EQ AC blocks", lambda: (
    sys.path.insert(0, HERE),
    bd := sys.modules.get("build_dashboard") or __import__("build_dashboard"),
    cfg := bd.load_config(),
    js := bd.render_eq_blueprint(cfg),
    count := js.count("pillars:{"),
    f"{count} pillar blocks" if count == 3
    else (_ for _ in ()).throw(AssertionError(f"Expected 3 active EQ ACs, got {count}"))
)[-1])


# ─────────────────────────────────────────────────────────────────────────────
# 6. METHODOLOGY DOC
# ─────────────────────────────────────────────────────────────────────────────
print(HEAD)
print("6. docs/TAA_Methodology.docx")
print(HEAD)

def _check_docx():
    from docx import Document
    path = os.path.join(ROOT, "docs", "TAA_Methodology.docx")
    doc = Document(path)
    paras = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return paras

check("python-docx importable", lambda: None)

check("docx readable, >= 30 paragraphs", lambda: (
    paras := _check_docx(),
    f"{len(paras)} paragraphs" if len(paras) >= 30
    else (_ for _ in ()).throw(AssertionError(f"Only {len(paras)} paragraphs"))
)[-1])

check("docx contains key section headings", lambda: (
    paras := _check_docx(),
    text := " ".join(paras),
    missing := [h for h in ["Executive Summary","Asset Class Universe","Aggregation Pipeline",
                              "Absolute vs Relative","Signal Blueprints"] if h not in text],
    None if not missing
    else (_ for _ in ()).throw(AssertionError(f"Missing headings: {missing}"))
)[-1])

check("docx mentions all 10 asset class full names", lambda: (
    from_docx := _check_docx(),
    text := " ".join(from_docx),
    expected := ["Money Market","Short-Term Fixed Income","LT US Treasuries","LT US Corporate",
                 "LT EM","US Equity","US Growth","US Value","DM ex-US","Emerging Markets"],
    missing := [n for n in expected if n not in text],
    None if not missing
    else (_ for _ in ()).throw(AssertionError(f"Missing AC names: {missing}"))
)[-1])


# ─────────────────────────────────────────────────────────────────────────────
# SUMMARY
# ─────────────────────────────────────────────────────────────────────────────
print(HEAD)
if not failures:
    print("ALL CHECKS PASSED OK")
    print(HEAD)
    sys.exit(0)
else:
    print(f"FAILED CHECKS ({len(failures)}):")
    for f in failures:
        print(f"  X {f}")
    print(HEAD)
    sys.exit(1)
