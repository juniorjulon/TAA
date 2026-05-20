"""
src/signal_engine.py
====================
Generic signal loading and normalisation engine.

Reads DataSeries from config/taa_config.xlsx. For every active row:
  1. Loads the raw series from the correct source
     - series_type="original"  -> Dashboard_TAA_Inputs.xlsx (any sheet by name)
     - series_type="custom"    -> data/custom_series.xlsx
  2. Applies the transform_code to normalise to z-score scale
  3. Returns a dict {series_id: pd.Series (z-score)}

Transform codes (all single-column, all output z-score scale):
  ewma_z    : EWMA z-score (span=window)
  rolling_z : Rolling z-score (window days)
  pctile    : Percentile rank 0-1, rescaled (p-0.5)*4 -> z-equivalent
  mom_z     : pct_change(window) -> ewma_zscore (EPS revisions)
  price_mom : Composite multi-horizon price momentum (MomentumConfig)
  inv_mom_z : -ewma_zscore(diff(window)); falling/tightening=positive

Usage:
    engine  = SignalEngine(config_xlsx, inputs_xlsx, custom_xlsx)
    signals = engine.load_all()   # {series_id: pd.Series}
"""

import os
import sys
import warnings
import pandas as pd
import numpy as np

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from config import CONFIG_XLSX, EXCEL_PATH, CUSTOM_SERIES_PATH, EWMA_SPAN, WINDOWS, MIN_DATE_FOR_SIGNALS
from signals import (ewma_zscore, rolling_zscore, pctile_rank,
                     composite_price_momentum)
from data_loader import load_raw_sheet


# ─────────────────────────────────────────────────────────────────────────────
# Percentile rescaling
# ─────────────────────────────────────────────────────────────────────────────

def pctile_to_z(p: pd.Series) -> pd.Series:
    """Rescale percentile [0,1] to z-score equivalent: (p - 0.5) * 4."""
    return ((p - 0.5) * 4).clip(-3, 3).rename(p.name)


# ─────────────────────────────────────────────────────────────────────────────
# Window parsing
# ─────────────────────────────────────────────────────────────────────────────

def _parse_window(raw) -> int:
    """Convert window string (e.g. '252', '756', '21') to int."""
    if raw is None:
        return EWMA_SPAN
    try:
        return max(21, int(str(raw).strip()))
    except (ValueError, TypeError):
        return EWMA_SPAN


# ─────────────────────────────────────────────────────────────────────────────
# MomentumConfig reader
# ─────────────────────────────────────────────────────────────────────────────

def _load_momentum_config(wb) -> dict:
    """
    Load MomentumConfig sheet from taa_config.xlsx.

    Returns dict: {series_id: dict of weights and window params}
    """
    if "MomentumConfig" not in wb.sheetnames:
        return {}

    ws = wb["MomentumConfig"]
    rows = list(ws.values)
    if not rows:
        return {}

    headers = [str(h).strip() for h in rows[0]]
    config = {}
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        d = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        sid = str(d.get("series_id", "")).strip()
        if sid:
            config[sid] = d
    return config


# ─────────────────────────────────────────────────────────────────────────────
# DataSeries loader
# ─────────────────────────────────────────────────────────────────────────────

def _load_dataseries(wb) -> list[dict]:
    """Load DataSeries sheet into list of dicts (one per row)."""
    ws = wb["DataSeries"]
    rows = list(ws.values)
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else f"col_{i}"
               for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        d = {}
        for i, h in enumerate(headers):
            d[h] = row[i] if i < len(row) else None
        result.append(d)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Transform functions
# ─────────────────────────────────────────────────────────────────────────────

def _apply_transform(raw: pd.Series, transform_code: str,
                     window: int, series_id: str,
                     mom_config: dict) -> pd.Series:
    """Apply the named transform to raw series; return z-score scale output."""
    s = raw.dropna()
    if s.empty:
        return pd.Series(dtype=float, name=series_id)

    tc = str(transform_code).strip().lower() if transform_code else "ewma_z"

    if tc == "ewma_z":
        return ewma_zscore(s, span=window).rename(series_id)

    elif tc == "rolling_z":
        return rolling_zscore(s, window=window).rename(series_id)

    elif tc == "pctile":
        p = pctile_rank(s, window=window)
        return pctile_to_z(p).rename(series_id)

    elif tc == "mom_z":
        chg = s.pct_change(window)
        return ewma_zscore(chg, span=EWMA_SPAN).rename(series_id)

    elif tc == "price_mom":
        cfg = mom_config.get(series_id, {})
        if cfg:
            return _composite_price_mom(s, cfg, series_id)
        return composite_price_momentum(s).rename(series_id)

    elif tc == "diff_z":
        # Direction-neutral: ewma_z(diff(window)).
        # Sign lives in SignalMapping — use sign=-1 for assets where falling=bullish
        # (credit spreads tightening, yields falling).
        diff = s.diff(window)
        return ewma_zscore(diff, span=EWMA_SPAN).rename(series_id)

    elif tc == "inv_mom_z":
        # DEPRECATED — kept for backward compatibility only.
        # Use diff_z + sign=-1 in SignalMapping instead (same math, cleaner design).
        warnings.warn(
            f"[signal_engine] '{series_id}' uses deprecated transform 'inv_mom_z'. "
            f"Replace with 'diff_z' in DataSeries and set sign=-1 in SignalMapping.",
            DeprecationWarning, stacklevel=2
        )
        diff = s.diff(window)
        return ewma_zscore(-diff, span=EWMA_SPAN).rename(series_id)

    else:
        warnings.warn(f"[signal_engine] Unknown transform_code '{tc}' for {series_id}; "
                      f"falling back to ewma_z")
        return ewma_zscore(s, span=window).rename(series_id)


def _composite_price_mom(price: pd.Series, cfg: dict, series_id: str) -> pd.Series:
    """
    Composite momentum using per-series MomentumConfig weights.

    Falls back to composite_price_momentum() default weights if MomentumConfig is missing
    or weights are all zero.
    """
    def _fw(key, default=0.0):
        v = cfg.get(key)
        try:
            return float(v) if v is not None else default
        except (ValueError, TypeError):
            return default

    def _iw(key, default=21):
        v = cfg.get(key)
        try:
            return max(1, int(v)) if v is not None else default
        except (ValueError, TypeError):
            return default

    w_1m   = _fw("ret_1m_weight",    0.10)
    w_3m   = _fw("ret_3m_weight",    0.25)
    w_12m  = _fw("ret_12_1m_weight", 0.40)
    w_ma   = _fw("ma_dist_weight",   0.15)
    w_rsi  = _fw("rsi_weight",       0.10)

    d_1m   = _iw("ret_1m_days",   21)
    d_3m   = _iw("ret_3m_days",   63)
    d_12m  = _iw("ret_12m_days",  252)
    d_skip = _iw("skip_days",     21)
    ma_f   = _iw("ma_fast",       50)
    ma_s   = _iw("ma_slow",       200)
    rsi_p  = _iw("rsi_period",    14)
    min_h  = 63

    signals = {}

    if len(price.dropna()) >= d_1m + min_h:
        signals["1m"] = (w_1m, ewma_zscore(price.pct_change(d_1m)))

    if len(price.dropna()) >= d_3m + min_h:
        signals["3m"] = (w_3m, ewma_zscore(price.pct_change(d_3m)))

    if len(price.dropna()) >= d_12m + d_skip + min_h:
        mom_12_1 = price.shift(d_skip).pct_change(d_12m - d_skip)
        signals["12_1m"] = (w_12m, ewma_zscore(mom_12_1))

    if len(price.dropna()) >= ma_s + min_h:
        ma_fast_s  = price.rolling(ma_f,  min_periods=ma_f  // 2).mean()
        ma_slow_s  = price.rolling(ma_s,  min_periods=ma_s  // 2).mean()
        signals["ma"] = (w_ma, ewma_zscore((ma_fast_s - ma_slow_s) / ma_slow_s.replace(0, np.nan)))

    if len(price.dropna()) >= rsi_p + min_h:
        delta  = price.diff()
        up     = delta.clip(lower=0)
        down   = (-delta).clip(lower=0)
        rs     = (up.ewm(span=rsi_p, adjust=False).mean() /
                  down.ewm(span=rsi_p, adjust=False).mean())
        rsi    = 100 - (100 / (1 + rs))
        signals["rsi"] = (w_rsi, ewma_zscore((rsi - 50) / 50))

    if not signals:
        return composite_price_momentum(price).rename(series_id)

    total_w = sum(w for w, _ in signals.values())
    if total_w < 1e-9:
        return composite_price_momentum(price).rename(series_id)

    comp = sum((w / total_w) * s for w, s in signals.values())
    return comp.clip(-3, 3).rename(series_id)


# ─────────────────────────────────────────────────────────────────────────────
# Sheet cache (lazy-loads sheets from Dashboard_TAA_Inputs.xlsx)
# ─────────────────────────────────────────────────────────────────────────────

class _SheetCache:
    def __init__(self, inputs_path: str, custom_path: str):
        self._inputs_path = inputs_path
        self._custom_path = custom_path
        self._cache: dict[str, pd.DataFrame] = {}

    def get_sheet(self, sheet_name: str) -> pd.DataFrame:
        if sheet_name in self._cache:
            return self._cache[sheet_name]
        if sheet_name == "custom_series":
            df = self._load_custom()
        else:
            try:
                df = load_raw_sheet(self._inputs_path, sheet_name)
            except Exception as e:
                warnings.warn(f"[signal_engine] Cannot load sheet '{sheet_name}': {e}")
                df = pd.DataFrame()
        self._cache[sheet_name] = df
        return df

    def _load_custom(self) -> pd.DataFrame:
        if not os.path.exists(self._custom_path):
            warnings.warn(f"[signal_engine] custom_series.xlsx not found: {self._custom_path}. "
                          f"Run python src/build_custom_series.py first.")
            return pd.DataFrame()
        df = pd.read_excel(self._custom_path, sheet_name="custom_series",
                           index_col=0, parse_dates=True)
        df.index = pd.to_datetime(df.index, errors="coerce")
        df = df[df.index.notna()].sort_index()
        return df


# ─────────────────────────────────────────────────────────────────────────────
# SignalEngine
# ─────────────────────────────────────────────────────────────────────────────

class SignalEngine:
    """
    Generic TAA signal loader and normaliser.

    Parameters
    ----------
    config_path  : path to config/taa_config.xlsx
    inputs_path  : path to data/Dashboard_TAA_Inputs.xlsx
    custom_path  : path to data/custom_series.xlsx
    """

    def __init__(self, config_path: str = None, inputs_path: str = None,
                 custom_path: str = None):
        self._config_path  = config_path  or CONFIG_XLSX
        self._inputs_path  = inputs_path  or EXCEL_PATH
        self._custom_path  = custom_path  or CUSTOM_SERIES_PATH
        self._wb           = None
        self._dataseries   = None
        self._mom_config   = None
        self._cache        = None

    def _ensure_loaded(self):
        if self._wb is None:
            self._wb         = load_workbook(self._config_path, data_only=True)
            self._dataseries = _load_dataseries(self._wb)
            self._mom_config = _load_momentum_config(self._wb)
            self._cache      = _SheetCache(self._inputs_path, self._custom_path)

    def load_all(self, verbose: bool = False) -> dict[str, pd.Series]:
        """
        Load and normalise all active series from DataSeries.

        Returns
        -------
        dict  {series_id: pd.Series (z-score, DatetimeIndex)}
        """
        self._ensure_loaded()
        signals: dict[str, pd.Series] = {}

        for row in self._dataseries:
            sid          = str(row.get("series_id", "") or "").strip()
            series_type  = str(row.get("series_type",  "") or "").strip().lower()
            input_sheet  = str(row.get("input_sheet",  "") or "").strip()
            input_column = str(row.get("input_column", "") or "").strip()
            transform_code = str(row.get("transform_code", "") or "").strip()
            window_raw   = row.get("window", None)

            if not sid:
                continue

            # Skip series without executable configuration
            if not series_type or series_type == "unavailable":
                continue
            if not input_sheet or input_sheet.lower() == "unavailable":
                continue
            if not input_column:
                continue
            if series_type not in ("original", "custom"):
                continue

            # Skip duplicates — keep first occurrence
            if sid in signals:
                continue

            window = _parse_window(window_raw)

            try:
                raw = self._load_raw(input_sheet, input_column)
                if raw.dropna().empty:
                    if verbose:
                        print(f"  EMPTY  {sid}")
                    continue

                z = _apply_transform(raw, transform_code, window, sid,
                                     self._mom_config)
                # Enforce global signal floor: drop unreliable EWMA warm-up period
                if MIN_DATE_FOR_SIGNALS:
                    z = z[z.index >= MIN_DATE_FOR_SIGNALS]
                if not z.dropna().empty:
                    signals[sid] = z
                    if verbose:
                        print(f"  OK     {sid:<28} {len(z.dropna()):>5} pts  [{transform_code}]")
                else:
                    if verbose:
                        print(f"  BLANK  {sid}")

            except Exception as exc:
                warnings.warn(f"[signal_engine] {sid}: {exc}")
                if verbose:
                    import traceback
                    traceback.print_exc()

        return signals

    def _load_raw(self, sheet_name: str, column: str) -> pd.Series:
        """Load a single column from the appropriate source."""
        df = self._cache.get_sheet(sheet_name)
        if df.empty:
            return pd.Series(dtype=float, name=column)
        if column not in df.columns:
            raise KeyError(f"Column '{column}' not found in sheet '{sheet_name}'. "
                           f"Available: {list(df.columns[:10])}")
        return df[column]

    def get_signal_mapping(self) -> pd.DataFrame:
        """Load SignalMapping sheet as a DataFrame for use in build_all_pillars()."""
        self._ensure_loaded()
        ws = self._wb["SignalMapping"]
        rows = list(ws.values)
        if not rows:
            return pd.DataFrame()
        headers = [str(h).strip() for h in rows[0]]
        data = [dict(zip(headers, row)) for row in rows[1:] if row and row[0]]
        return pd.DataFrame(data)
