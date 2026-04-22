"""
build_dashboard.py
==================
Reads `config/taa_config.xlsx` (user-owned source of truth) and regenerates the
machine-managed blocks in `index.html` and `src/config.py` between the
`<<<BUILD:...>>>` markers. Human-authored code outside the markers is preserved.

Flow:
    User edits Excel  ->  python src/build_dashboard.py  ->  index.html + src/config.py updated

Blocks regenerated in index.html:
    <<<BUILD:SIG_MATRIX>>>       pivot of every signal x every AC (signs)
    <<<BUILD:AC_META>>>          AC_ORDER + AC_SHORT
    <<<BUILD:FI_BLUEPRINT>>>     methodology cards for FI ACs
    <<<BUILD:EQ_BLUEPRINT>>>     methodology cards for EQ ACs
    <<<BUILD:AC_LABEL_PW>>>      AC_LABEL_FULL + PW (pillar weights)

Blocks regenerated in src/config.py:
    <<<BUILD:PY_AC_UNIVERSE>>>   ASSET_CLASSES + LABELS + GROUPS
    <<<BUILD:PY_PILLAR_WEIGHTS>>> PILLAR_WEIGHTS dict
    <<<BUILD:PY_MAX_TILT>>>      MAX_TILT_PCT dict
"""

from __future__ import annotations
import os
import re
from collections import defaultdict, OrderedDict
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "config", "taa_config.xlsx")
INDEX_HTML = os.path.join(ROOT, "index.html")
CONFIG_PY  = os.path.join(HERE, "config.py")

PILLAR_ORDER = ("F", "M", "S", "V")
PILLAR_NAME = {"F": "Fundamentals", "M": "Momentum", "S": "Sentiment", "V": "Valuation"}


# ─────────────────────────────────────────────────────────────────────────────
# READ EXCEL
# ─────────────────────────────────────────────────────────────────────────────
def _read_sheet(ws):
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append(dict(zip(headers, r)))
    return rows


def load_config():
    wb = load_workbook(XLSX, data_only=True)
    acs        = _read_sheet(wb["AssetClasses"])
    series     = _read_sheet(wb["DataSeries"])
    weights    = _read_sheet(wb["PillarWeights"])
    notes      = _read_sheet(wb["PillarNotes"])
    mapping    = _read_sheet(wb["SignalMapping"])
    return {
        "asset_classes":  acs,
        "data_series":    {s["series_id"]: s for s in series},
        "pillar_weights": {w["ac_id"]: w for w in weights},
        "pillar_notes":   {(n["ac_id"], n["pillar"]): n["note"] for n in notes},
        "mapping":        mapping,
    }


# ─────────────────────────────────────────────────────────────────────────────
# JS ESCAPING
# ─────────────────────────────────────────────────────────────────────────────
_ESC = str.maketrans({"\\": "\\\\", "'": "\\'", "\n": "\\n", "\r": "\\r"})


def js(s):
    if s is None:
        return "''"
    return "'" + str(s).translate(_ESC) + "'"


def pct(v):
    if v is None or v == "":
        return ""
    if isinstance(v, str):
        return v if v.endswith("%") else f"{v}%"
    return f"{int(round(float(v) * 100))}%" if float(v) <= 1 else f"{int(round(float(v)))}%"


# ─────────────────────────────────────────────────────────────────────────────
# RENDER JS BLOCKS
# ─────────────────────────────────────────────────────────────────────────────
def render_ac_meta(cfg):
    acs = cfg["asset_classes"]
    order = [a["ac_id"] for a in acs]
    short = {a["ac_id"]: a["short_label"] for a in acs}
    out = ["const AC_ORDER=[" + ",".join(js(a) for a in order) + "];"]
    pairs = ",".join(f"{a}:{js(short[a])}" for a in order)
    out.append("const AC_SHORT={" + pairs + "};")
    return "\n".join(out)


def render_ac_label_pw(cfg):
    acs = cfg["asset_classes"]
    pw  = cfg["pillar_weights"]
    lbl_pairs = ",".join(f"{a['ac_id']}:{js(a['full_label'])}" for a in acs)
    out = ["const AC_LABEL_FULL={" + lbl_pairs + "};"]
    def pw_obj(ac_id):
        w = pw[ac_id]
        return f"{ac_id}:{{F:{float(w['F']):.2f},M:{float(w['M']):.2f},S:{float(w['S']):.2f},V:{float(w['V']):.2f}}}"
    pw_pairs = ",".join(pw_obj(a["ac_id"]) for a in acs)
    out.append("const PW={" + pw_pairs + "};")
    return "\n".join(out)


def render_sig_matrix(cfg):
    """
    Aggregate SignalMapping by series_id and build one row per signal with a
    dict of {ac_id: sign}. Missing ACs default to '—'.
    """
    acs = [a["ac_id"] for a in cfg["asset_classes"]]
    series = cfg["data_series"]
    by_series = OrderedDict()
    for m in cfg["mapping"]:
        sid = m["series_id"]
        if sid not in series:
            continue
        if sid not in by_series:
            by_series[sid] = {}
        by_series[sid][m["ac_id"]] = m["sign"] or "—"

    # Order rows by pillar (F,M,S,V) then by series_id appearance in DataSeries
    series_order = list(series.keys())
    ordered_sids = sorted(
        by_series.keys(),
        key=lambda s: (PILLAR_ORDER.index(series[s]["pillar"]) if series[s]["pillar"] in PILLAR_ORDER else 9,
                       series_order.index(s)),
    )

    lines = ["const SIG_MATRIX=["]
    for sid in ordered_sids:
        s = series[sid]
        signs = by_series[sid]
        sign_items = ",".join(f"{ac}:{js(signs.get(ac, '—'))}" for ac in acs)
        lines.append(
            f"  {{pillar:{js(s['pillar'])},name:{js(s['signal_name'])},"
            f"source:{js(s['source'])},freq:{js(s['frequency'])},"
            f"signs:{{{sign_items}}}}},"
        )
    lines.append("];")
    return "\n".join(lines)


def _ac_block_id(ac):
    """Short alnum id for the blueprint block (stripped short_label)."""
    return re.sub(r"[^A-Za-z0-9]", "", ac["short_label"]) or ac["ac_id"]


def _pillar_pct_label(pillar_letter, pw_row):
    w = float(pw_row[pillar_letter])
    return f"{PILLAR_NAME[pillar_letter]} ({int(round(w * 100))}%)"


def _render_blueprint(cfg, group):
    """
    Build FI_BLUEPRINT or EQ_BLUEPRINT JS literal for the given group ('FI' or 'EQ').
    """
    acs = [a for a in cfg["asset_classes"] if a["group"] == group]
    series = cfg["data_series"]
    pw = cfg["pillar_weights"]
    notes = cfg["pillar_notes"]

    # Index SignalMapping by (ac_id, pillar) preserving original row order for weight display
    by_ac_pil = defaultdict(list)
    for m in cfg["mapping"]:
        by_ac_pil[(m["ac_id"], m["pillar"])].append(m)

    lines = []
    for a in acs:
        ac_id = a["ac_id"]
        bid   = _ac_block_id(a)
        header = (
            f"  {{id:{js(bid)}, name:{js(a['full_label'])}, "
            f"sub:{js(a['sub_description'])}, color:{js(a['color'])},"
        )
        lines.append(header)
        lines.append("   pillars:{")
        for p in PILLAR_ORDER:
            title = _pillar_pct_label(p, pw[ac_id])
            note  = notes.get((ac_id, p), "")
            sig_rows = by_ac_pil.get((ac_id, p), [])
            note_js = f"note:{js(note)}, " if note else ""
            lines.append(f"     {p}:{{title:{js(title)}, {note_js}signals:[")
            for m in sig_rows:
                sid = m["series_id"]
                s_meta = series.get(sid, {})
                name = s_meta.get("signal_name", sid)
                desc = m.get("description_override") or s_meta.get("notes") or ""
                sign = m.get("sign") or "—"
                wt   = pct(m.get("weight_in_pillar"))
                lines.append(
                    f"       {{n:{js(name)}, d:{js(desc)}, s:{js(sign)}, w:{js(wt)}}},"
                )
            lines.append("     ]},")
        lines.append("   }},")

    var = "FI_BLUEPRINT" if group == "FI" else "EQ_BLUEPRINT"
    return f"const {var} = [\n" + "\n".join(lines) + "\n];"


def render_fi_blueprint(cfg):
    return _render_blueprint(cfg, "FI")


def render_eq_blueprint(cfg):
    return _render_blueprint(cfg, "EQ")


# ─────────────────────────────────────────────────────────────────────────────
# RENDER PYTHON BLOCKS (for src/config.py)
# ─────────────────────────────────────────────────────────────────────────────
def render_py_ac_universe(cfg):
    acs = cfg["asset_classes"]
    ids = [a["ac_id"] for a in acs]
    out = ["ASSET_CLASSES = ["]
    for i, aid in enumerate(ids):
        comma = "," if i < len(ids) - 1 else ","
        out.append(f'    "{aid}"{comma}')
    out.append("]")
    out.append("")
    out.append("ASSET_CLASS_LABELS = {")
    for a in acs:
        out.append(f'    "{a["ac_id"]}": "{a["full_label"]}",')
    out.append("}")
    out.append("")
    out.append("ASSET_CLASS_GROUPS = {")
    for a in acs:
        out.append(f'    "{a["ac_id"]}": "{a["group"]}",')
    out.append("}")
    return "\n".join(out)


def render_py_pillar_weights(cfg):
    pw = cfg["pillar_weights"]
    out = ["PILLAR_WEIGHTS = {"]
    for a in cfg["asset_classes"]:
        w = pw[a["ac_id"]]
        out.append(
            f'    "{a["ac_id"]}": {{"F": {float(w["F"]):.2f}, "M": {float(w["M"]):.2f}, '
            f'"S": {float(w["S"]):.2f}, "V": {float(w["V"]):.2f}}},'
        )
    out.append("}")
    return "\n".join(out)


def render_py_max_tilt(cfg):
    out = ["MAX_TILT_PCT = {"]
    for a in cfg["asset_classes"]:
        out.append(f'    "{a["ac_id"]}": {float(a["max_tilt_pct"]):.1f},')
    out.append("}")
    return "\n".join(out)


# ─────────────────────────────────────────────────────────────────────────────
# MARKER REPLACEMENT
# ─────────────────────────────────────────────────────────────────────────────
def replace_block(text, start_marker, end_marker, new_body, comment_prefix="//"):
    """
    Replace the content between `<<<BUILD:X_START>>>` and `<<<BUILD:X_END>>>`
    with new_body. The marker lines themselves are preserved.
    """
    pattern = re.compile(
        rf"({re.escape(comment_prefix)}\s*<<<BUILD:{start_marker}>>>[^\n]*\n)"
        rf".*?"
        rf"(\n{re.escape(comment_prefix)}\s*<<<BUILD:{end_marker}>>>)",
        re.DOTALL,
    )
    replacement = r"\1" + new_body + r"\2"
    new_text, n = pattern.subn(replacement, text)
    if n == 0:
        raise RuntimeError(f"Markers {start_marker}/{end_marker} not found (prefix {comment_prefix!r})")
    return new_text


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    if not os.path.isfile(XLSX):
        raise SystemExit(f"Config Excel not found: {XLSX}\nRun `python src/seed_taa_config.py` first.")

    cfg = load_config()

    # ---- index.html ---------------------------------------------------------
    with open(INDEX_HTML, "r", encoding="utf-8") as f:
        html = f.read()

    replacements_html = [
        ("SIG_MATRIX_START",   "SIG_MATRIX_END",   render_sig_matrix(cfg)),
        ("AC_META_START",      "AC_META_END",      render_ac_meta(cfg)),
        ("FI_BLUEPRINT_START", "FI_BLUEPRINT_END", render_fi_blueprint(cfg)),
        ("EQ_BLUEPRINT_START", "EQ_BLUEPRINT_END", render_eq_blueprint(cfg)),
        ("AC_LABEL_PW_START",  "AC_LABEL_PW_END",  render_ac_label_pw(cfg)),
    ]
    for start, end, body in replacements_html:
        html = replace_block(html, start, end, body, comment_prefix="//")

    with open(INDEX_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] index.html regenerated ({len(replacements_html)} blocks)")

    # ---- src/config.py ------------------------------------------------------
    with open(CONFIG_PY, "r", encoding="utf-8") as f:
        py = f.read()

    py_blocks = [
        ("PY_AC_UNIVERSE_START",  "PY_AC_UNIVERSE_END",  render_py_ac_universe(cfg)),
        ("PY_PILLAR_WEIGHTS_START","PY_PILLAR_WEIGHTS_END", render_py_pillar_weights(cfg)),
        ("PY_MAX_TILT_START",     "PY_MAX_TILT_END",     render_py_max_tilt(cfg)),
    ]
    missing = []
    for start, end, body in py_blocks:
        if f"<<<BUILD:{start}>>>" not in py:
            missing.append(start)
            continue
        py = replace_block(py, start, end, body, comment_prefix="#")
    if missing:
        print(f"[WARN] markers not yet in config.py; skipped: {missing}")
    else:
        with open(CONFIG_PY, "w", encoding="utf-8") as f:
            f.write(py)
        print(f"[OK] src/config.py regenerated ({len(py_blocks)} blocks)")

    print("[DONE] Rebuild complete.")


if __name__ == "__main__":
    main()
