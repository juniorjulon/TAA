"""
src/extend_taa_config.py
========================
Extends config/taa_config.xlsx with the executable signal engine columns.

Run once (or re-run to update):
    python src/extend_taa_config.py

Changes made:
  1. Adds columns series_type, input_sheet, input_column, transform_code to DataSeries
  2. Appends new executable DataSeries rows (original + custom series)
  3. Appends new SignalMapping rows for all 12 ACs (executable wiring)
  4. Creates TransformCodes sheet (if not present)
  5. Creates MomentumConfig sheet (if not present)

Existing documentation rows in DataSeries and SignalMapping are preserved.
New rows use the execution series_ids that signal_engine.py references.
"""

from __future__ import annotations
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "config", "taa_config.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1E2E47")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
THIN        = Side(border_style="thin", color="D1D5DB")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GREEN_FILL  = PatternFill("solid", fgColor="D1FAE5")
BLUE_FILL   = PatternFill("solid", fgColor="DBEAFE")


def _hdr(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER


def _autosize(ws, min_w=10, max_w=55):
    for col in ws.columns:
        letter = col[0].column_letter
        longest = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[letter].width = max(min_w, min(max_w, longest + 2))


# ─────────────────────────────────────────────────────────────────────────────
# 1 — NEW DataSeries ROWS (executable catalog)
# ─────────────────────────────────────────────────────────────────────────────
# Format: (series_id, signal_name, ticker, source, frequency, pillar,
#          transformation, window, notes,
#          series_type, input_sheet, input_column, transform_code)

EXEC_ORIGINAL = [
    # ── CESI ──────────────────────────────────────────────────────────────────
    ("cesi_us",       "Citi US Eco Surprise",    "CESIUSD Index",         "BBG", "Daily",   "F", "EWMA z-score", "252",  "Contrarian at pctile>85/<15", "original", "H1", "CESIUSD Index",         "ewma_z"),
    ("cesi_ez",       "Citi EZ Eco Surprise",    "CESIEUR Index",         "BBG", "Daily",   "F", "EWMA z-score", "252",  "European data surprises",     "original", "H1", "CESIEUR Index",         "ewma_z"),
    ("cesi_china",    "Citi China Eco Surprise", "CESICNY Index",         "BBG", "Daily",   "F", "EWMA z-score", "252",  "China macro beat/miss",       "original", "H1", "CESICNY Index",         "ewma_z"),
    ("cesi_em",       "Citi EM Eco Surprise",    "CESIEM Index",          "BBG", "Daily",   "F", "EWMA z-score", "252",  "Broad EM surprises",          "original", "H2", "CESIEM Index",          "ewma_z"),
    ("cesi_japan",    "Citi Japan Eco Surprise", "CESIJPY Index",         "BBG", "Daily",   "F", "EWMA z-score", "252",  "Japan macro surprises",       "original", "H2", "CESIJPY Index",         "ewma_z"),
    # ── Fundamentals: single-column originals ─────────────────────────────────
    ("pmi_japan_mfg", "Japan Mfg PMI",           "MPMIJPMA Index",        "BBG", "Monthly", "F", "EWMA z-score", "252",  "Japanese industrial PMI",     "original", "H2", "MPMIJPMA Index",        "ewma_z"),
    ("eps_us",        "US Fwd EPS Revision",     "SPX Fwd EPS",           "BBG", "Daily",   "F", "mom_z",        "21",   "pct_change(21)→ewma_z",       "original", "H3", "SPX Index Forward EPS", "mom_z"),
    ("eps_em",        "EM Fwd EPS Revision",     "MXEM Fwd EPS",          "BBG", "Daily",   "F", "mom_z",        "21",   "pct_change(21)→ewma_z",       "original", "H3", "MXEM Index Forward EPS","mom_z"),
    ("eps_eafe",      "EAFE Fwd EPS Revision",   "MXEF Fwd EPS",          "BBG", "Daily",   "F", "mom_z",        "21",   "pct_change(21)→ewma_z",       "original", "H3", "MXEF Index Forward EPS","mom_z"),
    ("eps_china",     "China Fwd EPS Revision",  "MXCN Fwd EPS",          "BBG", "Daily",   "F", "mom_z",        "21",   "pct_change(21)→ewma_z",       "original", "H3", "MXCN Index Forward EPS","mom_z"),
    ("eps_japan",     "Japan Fwd EPS Revision",  "MXJP Fwd EPS",          "BBG", "Daily",   "F", "mom_z",        "21",   "pct_change(21)→ewma_z",       "original", "H3", "MXJP Index Forward EPS","mom_z"),
    ("breakeven_5y",  "5Y Breakeven Inflation",  "USGGBE05 Index",        "BBG", "Daily",   "F", "rolling_z",    "1260", "10Y window z-score",          "original", "H5", "USGGBE05 Index",        "rolling_z"),
    ("breakeven_10y", "10Y Breakeven Inflation", "USGGBE10 Index",        "BBG", "Daily",   "F", "rolling_z",    "1260", "10Y window z-score",          "original", "H5", "USGGBE10 Index",        "rolling_z"),
    ("core_pce",      "Core PCE YoY",            "PCE CYOY Index",        "BBG", "Monthly", "F", "ewma_z",       "756",  "Fed preferred inflation gauge","original","H5", "PCE CYOY Index",        "ewma_z"),
    # ── Sentiment: vol & macro ─────────────────────────────────────────────────
    ("vix",           "CBOE VIX Index",          "VIX Index",             "BBG", "Daily",   "S", "ewma_z",       "1260", "Contrarian equity; FQ for FI","original", "H5", "VIX Index",             "ewma_z"),
    ("move_z",        "ICE MOVE Index",          "MOVE Index",            "BBG", "Daily",   "S", "ewma_z",       "1260", "Bond vol; FQ signal for FI",  "original", "H5", "MOVE Index",            "ewma_z"),
    ("vstoxx_z",      "VSTOXX (EZ Equity Vol)",  "V2X Index",             "BBG", "Daily",   "S", "ewma_z",       "1260", "EZ equity contrarian",        "original", "H5", "V2X Index",             "ewma_z"),
    ("dxy_z",         "DXY USD Index",           "DXY Curncy",            "BBG", "Daily",   "S", "ewma_z",       "252",  "Strong USD = EM headwind (-1)","original","H5", "DXY Curncy",            "ewma_z"),
    ("fci_z",         "Bloomberg US FCI",        "BFCIUS Index",          "BBG", "Daily",   "S", "ewma_z",       "252",  "Tight FCI = US equity/credit headwind","original","H5","BFCIUS Index",    "ewma_z"),
    ("aaii_z",        "AAII Bull-Bear Spread",   "Bull-Bear Spread",      "AAII","Weekly",  "S", "ewma_z",       "1260", "Contrarian; euphoria=bearish","original","AAII","Bull-Bear Spread",      "ewma_z"),
    # ── Valuation: OAS levels (pctile) ────────────────────────────────────────
    ("oas_bbb",       "BBB OAS Valuation",       "BAMLC0A4CBBB",          "FRED","Daily",   "V", "pctile",       "1260", "Wide=cheap=positive; 5Y pctile","original","OAS","BAMLC0A4CBBB",          "pctile"),
    ("oas_hy",        "HY OAS Valuation",        "BAMLH0A0HYM2",          "FRED","Daily",   "V", "pctile",       "1260", "Wide=cheap=positive; 5Y pctile","original","OAS","BAMLH0A0HYM2",          "pctile"),
    ("oas_em",        "EM Corp OAS Valuation",   "BAMLEM2BRRBBBCRPIOAS",  "FRED","Daily",   "V", "pctile",       "1260", "Wide=cheap=positive",         "original", "OAS","BAMLEM2BRRBBBCRPIOAS",  "pctile"),
    ("oas_latam",     "LatAm Corp OAS",          "BAMLEMRLCRPILAOAS",     "FRED","Daily",   "V", "pctile",       "1260", "Wide=cheap=positive",         "original", "OAS","BAMLEMRLCRPILAOAS",     "pctile"),
    # ── Momentum: OAS (inv_mom_z) ─────────────────────────────────────────────
    ("oas_bbb_mom",   "BBB OAS Momentum",        "BAMLC0A4CBBB",          "FRED","Daily",   "M", "inv_mom_z",    "21",   "Tightening=positive; diff window","original","OAS","BAMLC0A4CBBB",       "inv_mom_z"),
    ("oas_hy_mom",    "HY OAS Momentum",         "BAMLH0A0HYM2",          "FRED","Daily",   "M", "inv_mom_z",    "21",   "Tightening=positive",         "original", "OAS","BAMLH0A0HYM2",          "inv_mom_z"),
    ("oas_em_mom",    "EM Corp OAS Momentum",    "BAMLEM2BRRBBBCRPIOAS",  "FRED","Daily",   "M", "inv_mom_z",    "21",   "Tightening=positive",         "original", "OAS","BAMLEM2BRRBBBCRPIOAS",  "inv_mom_z"),
    # ── Valuation: yield levels (pctile) ──────────────────────────────────────
    ("gt02",          "US 2Y Yield Level",       "GT02 @BGN Govt",        "BBG", "Daily",   "V", "pctile",       "2520", "High yield=attractive carry", "original", "H5", "GT02 @BGN Govt",        "pctile"),
    ("gt10",          "US 10Y Yield Level",      "GT10 @BGN Govt",        "BBG", "Daily",   "V", "pctile",       "2520", "High yield=attractive carry", "original", "H5", "GT10 @BGN Govt",        "pctile"),
    ("tips_5y",       "TIPS 5Y Real Yield",      "H15X5YR Index",         "BBG", "Daily",   "V", "pctile",       "2520", "High real yield=attractive",  "original", "H5", "H15X5YR Index",         "pctile"),
    ("tips_10y",      "TIPS 10Y Real Yield",     "H15X10YR Index",        "BBG", "Daily",   "V", "pctile",       "2520", "High real yield=attractive",  "original", "H5", "H15X10YR Index",        "pctile"),
    # ── Momentum: yield (inv_mom_z) ───────────────────────────────────────────
    ("gt02_mom",      "US 2Y Yield Momentum",    "GT02 @BGN Govt",        "BBG", "Daily",   "M", "inv_mom_z",    "21",   "Falling yield=bond price up=+","original","H5", "GT02 @BGN Govt",        "inv_mom_z"),
    ("gt10_mom",      "US 10Y Yield Momentum",   "GT10 @BGN Govt",        "BBG", "Daily",   "M", "inv_mom_z",    "21",   "Falling yield=bond price up=+","original","H5", "GT10 @BGN Govt",        "inv_mom_z"),
    # ── Momentum: TR price (price_mom) ────────────────────────────────────────
    ("sp500_tr",      "S&P 500 TR",              "SPXT Index TR",         "BBG", "Daily",   "M", "price_mom",    "252",  "12-1M+3M+MA+RSI composite",   "original", "H4", "SPXT Index TR",         "price_mom"),
    ("sp500_gro_tr",  "S&P 500 Growth TR",       "SPTRSGX Index TR",      "BBG", "Daily",   "M", "price_mom",    "252",  "Growth style momentum",       "original", "H4", "SPTRSGX Index TR",      "price_mom"),
    ("sp500_val_tr",  "S&P 500 Value TR",        "SPTRSVX Index TR",      "BBG", "Daily",   "M", "price_mom",    "252",  "Value style momentum",        "original", "H4", "SPTRSVX Index TR",      "price_mom"),
    ("eafe_tr",       "MSCI EAFE TR",            "NDDUEAFE Index TR",     "BBG", "Daily",   "M", "price_mom",    "252",  "DM ex-US momentum",           "original", "H4", "NDDUEAFE Index TR",     "price_mom"),
    ("msci_em_tr",    "MSCI EM TR",              "NDUEEGF Index TR",      "BBG", "Daily",   "M", "price_mom",    "252",  "EM equity momentum",          "original", "H4", "NDUEEGF Index TR",      "price_mom"),
    ("em_xchina_tr",  "MSCI EM ex-China TR",     "M1CXBRV Index TR",      "BBG", "Daily",   "M", "price_mom",    "252",  "EM ex-China momentum",        "original", "H4", "M1CXBRV Index TR",      "price_mom"),
    ("china_tr",      "MSCI China TR",           "NDEUCHF Index TR",      "BBG", "Daily",   "M", "price_mom",    "252",  "China equity momentum",       "original", "H4", "NDEUCHF Index TR",      "price_mom"),
    ("msci_acwi_tr",  "MSCI ACWI TR",            "NDUEACWF Index TR",     "BBG", "Daily",   "M", "price_mom",    "252",  "Global equity momentum",      "original", "H4", "NDUEACWF Index TR",     "price_mom"),
    ("bfu5_price",    "BFU5 / 1-3Y UST TR",      "I13282US Index TR",     "BBG", "Daily",   "M", "price_mom",    "252",  "Short-term FI momentum",      "original", "H4", "I13282US Index TR",     "price_mom"),
    ("i132_price",    "ICE 1-3Y UST TR",         "I13282US Index TR",     "BBG", "Daily",   "M", "price_mom",    "252",  "Short-term FI momentum (alt)","original", "H4", "I13282US Index TR",     "price_mom"),
    ("lt03_price",    "LT03 Short Govt TR",      "LT03TRUU Index TR",     "BBG", "Daily",   "M", "price_mom",    "252",  "MM/STFI bond TR momentum",    "original", "H4", "LT03TRUU Index TR",     "price_mom"),
    ("bsgv_price",    "BSGV Long Govt TR",       "BSGVTRUU Index TR",     "BBG", "Daily",   "M", "price_mom",    "252",  "Long UST TR momentum",        "original", "H4", "BSGVTRUU Index TR",     "price_mom"),
]

EXEC_CUSTOM = [
    # ── PMI composites ─────────────────────────────────────────────────────────
    ("pmi_us",         "PMI US Composite",        "(NAPMPMI+NAPMNMI)/2",   "BBG","Monthly","F","ewma_z",  "252",   "ISM Mfg+Svcs average",                "custom","custom_series","pmi_us",         "ewma_z"),
    ("pmi_ez",         "PMI Eurozone Composite",  "(MPMIEZMA+MPMIEZSA)/2", "BBG","Monthly","F","ewma_z",  "252",   "EZ Mfg+Svcs average",                 "custom","custom_series","pmi_ez",         "ewma_z"),
    ("pmi_china",      "PMI China Composite",     "(CPMINDX+MPMICNSA)/2",  "BBG","Monthly","F","ewma_z",  "252",   "Caixin Mfg+Svcs average",             "custom","custom_series","pmi_china",      "ewma_z"),
    # ── GDP blends ─────────────────────────────────────────────────────────────
    ("gdp_us",         "US GDP Blend",            "ECGDUS 26+27 blended",  "BBG","Daily",  "F","ewma_z",  "252",   "w_cur=month/12; revision signal",     "custom","custom_series","gdp_us",         "ewma_z"),
    ("gdp_dm",         "DM GDP Blend",            "ECGDD1 26+27 blended",  "BBG","Daily",  "F","ewma_z",  "252",   "Developed markets GDP blend",         "custom","custom_series","gdp_dm",         "ewma_z"),
    ("gdp_em",         "EM GDP Blend",            "ECGDM1 26+27 blended",  "BBG","Daily",  "F","ewma_z",  "252",   "EM consensus GDP blend",              "custom","custom_series","gdp_em",         "ewma_z"),
    ("gdp_eu",         "EU GDP Blend",            "ECGDEU 26+27 blended",  "BBG","Daily",  "F","ewma_z",  "252",   "Eurozone GDP blend",                  "custom","custom_series","gdp_eu",         "ewma_z"),
    ("gdp_china",      "China GDP Blend",         "ECGDCN 26+27 blended",  "BBG","Daily",  "F","ewma_z",  "252",   "China GDP consensus blend",           "custom","custom_series","gdp_china",      "ewma_z"),
    ("gdp_japan",      "Japan GDP Blend",         "ECGDJP 26+27 blended",  "BBG","Daily",  "F","ewma_z",  "252",   "Japan GDP consensus blend",           "custom","custom_series","gdp_japan",      "ewma_z"),
    # ── Rate environment ───────────────────────────────────────────────────────
    ("modern_ted",     "Modern TED Spread",       "GB03-SOFRRATE",         "BBG","Daily",  "S","ewma_z",  "756",   "3M T-bill minus SOFR; funding stress","custom","custom_series","modern_ted",     "ewma_z"),
    ("real_ff",        "Real Fed Funds Rate",     "FDTR-PCE(ffill35d)",    "BBG","Daily",  "F","rolling_z","756",  "FDTR-PCE_YoY; restrictive>0",         "custom","custom_series","real_ff",        "rolling_z"),
    ("term_spread",    "Term Spread 10Y-2Y",      "GT10-GT02",             "BBG","Daily",  "V","ewma_z",  "756",   "Inverted<0; signal varies by AC",     "custom","custom_series","term_spread",    "ewma_z"),
    # ── ERP ────────────────────────────────────────────────────────────────────
    ("erp_us",         "US ERP (EY-TIPS10Y)",     "sp500_ey-tips_10y",     "Calc","Daily", "V","rolling_z","2520", "EY%-TIPS%; >4%=cheap; <1%=expensive", "custom","custom_series","erp_us",         "rolling_z"),
    ("erp_acwi",       "ACWI ERP",                "acwi_ey-tips_10y",      "Calc","Daily", "V","rolling_z","2520", "Global ERP",                          "custom","custom_series","erp_acwi",       "rolling_z"),
    ("erp_em",         "EM ERP",                  "em_ey-tips_10y",        "Calc","Daily", "V","rolling_z","2520", "EM equity ERP",                       "custom","custom_series","erp_em",         "rolling_z"),
    ("erp_em_xchina",  "EM ex-China ERP",         "emx_ey-tips_10y",       "Calc","Daily", "V","rolling_z","2520", "EM ex-China ERP",                     "custom","custom_series","erp_em_xchina",  "rolling_z"),
    ("erp_china",      "China ERP",               "china_ey-tips_10y",     "Calc","Daily", "V","rolling_z","2520", "China equity ERP",                    "custom","custom_series","erp_china",      "rolling_z"),
    # ── PE scores ──────────────────────────────────────────────────────────────
    ("pe_score_sp500", "S&P 500 PE Score",        "pe_score(sp500_pe)",    "Calc","Daily", "V","ewma_z",  "756",   "-2+4*(1-pctile); cheap=positive",     "custom","custom_series","pe_score_sp500", "ewma_z"),
    ("pe_score_eafe",  "EAFE PE Score",           "pe_score(eafe_pe)",     "Calc","Daily", "V","ewma_z",  "756",   "DM ex-US valuation score",            "custom","custom_series","pe_score_eafe",  "ewma_z"),
    ("pe_score_em",    "MSCI EM PE Score",        "pe_score(em_pe)",       "Calc","Daily", "V","ewma_z",  "756",   "EM valuation score",                  "custom","custom_series","pe_score_em",    "ewma_z"),
    ("pe_score_emx",   "EM ex-China PE Score",    "pe_score(emx_pe)",      "Calc","Daily", "V","ewma_z",  "756",   "EM ex-China valuation score",         "custom","custom_series","pe_score_emx",   "ewma_z"),
    ("pe_score_china", "China PE Score",          "pe_score(china_pe)",    "Calc","Daily", "V","ewma_z",  "756",   "China valuation score",               "custom","custom_series","pe_score_china", "ewma_z"),
    ("pe_score_gro",   "US Growth PE Score",      "pe_score(growth_pe)",   "Calc","Daily", "V","ewma_z",  "756",   "Growth style valuation",              "custom","custom_series","pe_score_gro",   "ewma_z"),
    ("pe_score_val",   "US Value PE Score",       "pe_score(value_pe)",    "Calc","Daily", "V","ewma_z",  "756",   "Value style valuation",               "custom","custom_series","pe_score_val",   "ewma_z"),
    # ── Relative PE ────────────────────────────────────────────────────────────
    ("rel_pe_gro_val", "Growth vs Value Rel PE",  "log(PE_val/PE_gro)",    "Calc","Daily", "V","ewma_z",  "756",   "Growth cheap vs value=positive",      "custom","custom_series","rel_pe_gro_val", "ewma_z"),
    ("rel_pe_val_gro", "Value vs Growth Rel PE",  "log(PE_gro/PE_val)",    "Calc","Daily", "V","ewma_z",  "756",   "Value cheap vs growth=positive",      "custom","custom_series","rel_pe_val_gro", "ewma_z"),
    ("rel_pe_dm_us",   "DM vs US Rel PE",         "log(PE_sp500/PE_eafe)", "Calc","Daily", "V","ewma_z",  "756",   "DM cheap vs US=positive",             "custom","custom_series","rel_pe_dm_us",   "ewma_z"),
    ("rel_pe_em_us",   "EM vs US Rel PE",         "log(PE_sp500/PE_em)",   "Calc","Daily", "V","ewma_z",  "756",   "EM cheap vs US=positive",             "custom","custom_series","rel_pe_em_us",   "ewma_z"),
    ("rel_pe_us_em",   "US vs EM Rel PE",         "log(PE_em/PE_sp500)",   "Calc","Daily", "V","ewma_z",  "756",   "US cheap vs EM=positive",             "custom","custom_series","rel_pe_us_em",   "ewma_z"),
    ("rel_pe_china_us","China vs US Rel PE",      "log(PE_sp500/PE_china)","Calc","Daily", "V","ewma_z",  "756",   "China cheap vs US=positive",          "custom","custom_series","rel_pe_china_us","ewma_z"),
    # ── OAS-derived ────────────────────────────────────────────────────────────
    ("hy_ig_ratio",    "HY/IG Spread Ratio",      "oas_hy/oas_bbb",        "Calc","Daily", "V","ewma_z",  "756",   "Wide ratio=high risk premium=cheap",  "custom","custom_series","hy_ig_ratio",    "ewma_z"),
    ("hy_stress",      "HY Spread Stress Proxy",  "oas_hy.diff(21)",       "Calc","Daily", "S","ewma_z",  "756",   "+widening=stress; sign=-1 for credit","custom","custom_series","hy_stress",      "ewma_z"),
    ("em_stress",      "EM Spread Stress Proxy",  "oas_em.diff(21)",       "Calc","Daily", "S","ewma_z",  "756",   "+widening=EM stress; sign=-1 for EM", "custom","custom_series","em_stress",      "ewma_z"),
    ("hy_safe_haven",  "HY Widening=Safe Haven",  "oas_hy.diff(21)",       "Calc","Daily", "S","ewma_z",  "756",   "+widening=flight to safety; sign=+1", "custom","custom_series","hy_safe_haven",  "ewma_z"),
    ("embi",           "EMBI Proxy (EM OAS chg)", "oas_em.diff(21)",       "Calc","Daily", "S","ewma_z",  "756",   "+widening=EM stress; sign=-1 for EM", "custom","custom_series","embi",           "ewma_z"),
    # ── CDX momentum ───────────────────────────────────────────────────────────
    ("cdx_hy_mom",     "CDX HY Price Momentum",   "cdx_hy_momentum()",     "BBG","Daily",  "M","ewma_z",  "756",   "1M+3M blend; rising price=risk-on",   "custom","custom_series","cdx_hy_mom",     "ewma_z"),
    ("cdx_ig_mom",     "CDX IG Spread Momentum",  "cdx_ig_momentum()",     "BBG","Daily",  "M","ewma_z",  "756",   "1M+3M blend; tightening=positive",    "custom","custom_series","cdx_ig_mom",     "ewma_z"),
]

# ─────────────────────────────────────────────────────────────────────────────
# 2 — NEW SignalMapping ROWS
# ─────────────────────────────────────────────────────────────────────────────
# Format: (ac_id, series_id, pillar, sign, weight_in_pillar, description_override)
# Weight as string "35%" — consistent with existing format.

def pct(w: float) -> str:
    return f"{int(round(w * 100))}%"

SM_ROWS = []

def _sm(ac, sid, p, sign, w, desc=""):
    SM_ROWS.append((ac, sid, p, sign, pct(w), desc))

# ── MONEY MARKET ──────────────────────────────────────────────────────────────
ac = "money_market"
_sm(ac, "real_ff",       "F", "+1", 0.50, "Real FF > 0 = restrictive = MM attractive")
_sm(ac, "core_pce",      "F", "+1", 0.30, "High PCE = above target = restrictive")
_sm(ac, "breakeven_5y",  "F", "+1", 0.20, "Inflation elevated = restrictive policy")
_sm(ac, "lt03_price",    "M", "+1", 0.60, "Short govt TR momentum")
_sm(ac, "gt02_mom",      "M", "+1", 0.40, "2Y yield falling = bond prices rising")
_sm(ac, "modern_ted",    "S", "+1", 0.40, "TED wide = systemic stress = safe haven")
_sm(ac, "hy_safe_haven", "S", "+1", 0.30, "HY widening = flight to safety")
_sm(ac, "vix",           "S", "+1", 0.30, "High VIX = flight to quality")
_sm(ac, "gt02",          "V", "+1", 0.50, "2Y yield pctile; high = attractive carry")
_sm(ac, "term_spread",   "V", "-1", 0.25, "Inverted curve = high 2Y carry = bullish MM")
_sm(ac, "tips_5y",       "V", "+1", 0.25, "TIPS 5Y real yield; high = MM attractive")

# ── SHORT-TERM FI ─────────────────────────────────────────────────────────────
ac = "short_term_fi"
_sm(ac, "pmi_us",        "F", "-1", 0.30, "PMI up = growth = rates rise = bearish duration")
_sm(ac, "cesi_us",       "F", "-1", 0.25, "CESI up = growth surprise = bearish FI")
_sm(ac, "gdp_us",        "F", "-1", 0.20, "GDP revision up = rates rise")
_sm(ac, "core_pce",      "F", "-1", 0.15, "PCE above target = bearish duration")
_sm(ac, "breakeven_5y",  "F", "-1", 0.10, "Higher breakeven = bearish FI")
_sm(ac, "bfu5_price",    "M", "+1", 0.35, "1-3Y UST TR momentum")
_sm(ac, "i132_price",    "M", "+1", 0.20, "ICE 1-3Y UST momentum (alt)")
_sm(ac, "gt02_mom",      "M", "+1", 0.20, "2Y yield momentum (falling = positive)")
_sm(ac, "oas_bbb_mom",   "M", "+1", 0.15, "BBB spread tightening = positive for STFI")
_sm(ac, "cdx_ig_mom",    "M", "+1", 0.10, "CDX IG momentum")
_sm(ac, "move_z",        "S", "+1", 0.35, "Bond vol high = uncertainty = STFI safe haven")
_sm(ac, "vix",           "S", "+1", 0.35, "High VIX = flight to quality")
_sm(ac, "modern_ted",    "S", "+1", 0.30, "TED wide = funding stress = safe haven")
_sm(ac, "gt02",          "V", "+1", 0.35, "2Y yield level pctile; high = attractive")
_sm(ac, "oas_bbb",       "V", "+1", 0.25, "BBB OAS pctile; wide = attractive credit")
_sm(ac, "term_spread",   "V", "-1", 0.20, "Inverted curve = high 2Y carry = bullish STFI")
_sm(ac, "tips_5y",       "V", "+1", 0.20, "TIPS 5Y real yield attractiveness")

# ── LT TREASURIES ─────────────────────────────────────────────────────────────
ac = "lt_treasuries"
_sm(ac, "pmi_us",        "F", "-1", 0.30, "Growth = rates rise = bearish duration")
_sm(ac, "cesi_us",       "F", "-1", 0.25, "Surprise up = bearish duration")
_sm(ac, "gdp_us",        "F", "-1", 0.25, "GDP up = rate risk")
_sm(ac, "breakeven_10y", "F", "-1", 0.20, "10Y inflation = bearish long duration")
_sm(ac, "bsgv_price",    "M", "+1", 0.45, "Long govt TR momentum")
_sm(ac, "gt10_mom",      "M", "+1", 0.35, "10Y yield falling = LT bond prices up")
_sm(ac, "oas_bbb_mom",   "M", "+1", 0.20, "Credit momentum = risk-on confirms FI")
_sm(ac, "vix",           "S", "+1", 0.40, "High VIX = flight to quality LT UST")
_sm(ac, "move_z",        "S", "+1", 0.30, "Bond vol = safe haven demand")
_sm(ac, "modern_ted",    "S", "+1", 0.30, "Funding stress = safe haven LT bonds")
_sm(ac, "gt10",          "V", "+1", 0.35, "10Y yield pctile; high = attractive")
_sm(ac, "tips_10y",      "V", "+1", 0.30, "TIPS 10Y real yield attractiveness")
_sm(ac, "term_spread",   "V", "+1", 0.25, "Steep curve = attractive carry for LT bonds")
_sm(ac, "oas_bbb",       "V", "+1", 0.10, "Credit conditions proxy for LT bonds")

# ── LT US CORPORATE ───────────────────────────────────────────────────────────
ac = "lt_us_corp"
_sm(ac, "pmi_us",        "F", "+1", 0.25, "Growth = tighter credit spreads = positive")
_sm(ac, "cesi_us",       "F", "+1", 0.20, "Macro surprise = bullish credit")
_sm(ac, "gdp_us",        "F", "+1", 0.20, "GDP growth = lower default risk")
_sm(ac, "eps_us",        "F", "+1", 0.25, "Earnings revisions = corporate health")
_sm(ac, "breakeven_10y", "F", "-1", 0.10, "High inflation = duration risk for corp")
_sm(ac, "bfu5_price",    "M", "+1", 0.20, "Corporate TR momentum")
_sm(ac, "oas_bbb_mom",   "M", "+1", 0.30, "BBB spread tightening = direct signal")
_sm(ac, "oas_hy_mom",    "M", "+1", 0.20, "HY tightening = risk appetite")
_sm(ac, "cdx_ig_mom",    "M", "+1", 0.20, "CDX IG = live credit momentum")
_sm(ac, "gt10_mom",      "M", "+1", 0.10, "Rates falling = corp bond prices up")
_sm(ac, "vix",           "S", "-1", 0.30, "High VIX = risk off = bad for credit")
_sm(ac, "hy_stress",     "S", "-1", 0.25, "HY widening = credit stress = negative")
_sm(ac, "modern_ted",    "S", "-1", 0.20, "Funding stress = bad for credit")
_sm(ac, "cdx_hy_mom",    "S", "+1", 0.25, "CDX HY up = risk appetite = positive")
_sm(ac, "oas_bbb",       "V", "+1", 0.35, "BBB OAS pctile; wide = cheap credit")
_sm(ac, "oas_hy",        "V", "+1", 0.25, "HY OAS pctile; wide = cheap")
_sm(ac, "hy_ig_ratio",   "V", "+1", 0.20, "HY/IG ratio wide = attractive risk premium")
_sm(ac, "gt10",          "V", "+1", 0.20, "10Y yield pctile = underlying carry")

# ── LT EM FIXED INCOME ────────────────────────────────────────────────────────
ac = "lt_em_fi"
_sm(ac, "pmi_china",     "F", "+1", 0.30, "China PMI = EM growth engine")
_sm(ac, "cesi_em",       "F", "+1", 0.25, "EM macro surprises")
_sm(ac, "gdp_em",        "F", "+1", 0.25, "EM growth = spreads tighten")
_sm(ac, "eps_em",        "F", "+1", 0.20, "EM earnings trend = credit quality")
_sm(ac, "oas_em_mom",    "M", "+1", 0.40, "EM spread tightening = positive")
_sm(ac, "em_xchina_tr",  "M", "+1", 0.25, "EM ex-China equity momentum")
_sm(ac, "msci_em_tr",    "M", "+1", 0.35, "MSCI EM equity momentum")
_sm(ac, "dxy_z",         "S", "-1", 0.30, "Strong USD = EM headwind")
_sm(ac, "embi",          "S", "-1", 0.30, "EM sovereign stress = negative")
_sm(ac, "em_stress",     "S", "-1", 0.20, "EM credit stress = negative")
_sm(ac, "vix",           "S", "-1", 0.20, "High VIX = risk off = bad for EM FI")
_sm(ac, "oas_em",        "V", "+1", 0.45, "EM Corp OAS pctile; wide = cheap")
_sm(ac, "oas_latam",     "V", "+1", 0.25, "LatAm OAS pctile; wide = cheap")
_sm(ac, "gt10",          "V", "+1", 0.30, "US 10Y yield = base rate context for EM")

# ── US EQUITY ─────────────────────────────────────────────────────────────────
ac = "us_equity"
_sm(ac, "pmi_us",        "F", "+1", 0.35, "ISM composite = US growth engine")
_sm(ac, "cesi_us",       "F", "+1", 0.25, "Macro surprise = bullish equity")
_sm(ac, "gdp_us",        "F", "+1", 0.20, "GDP growth = earnings support")
_sm(ac, "eps_us",        "F", "+1", 0.15, "EPS revisions = direct earnings signal")
_sm(ac, "breakeven_5y",  "F", "-1", 0.05, "High inflation = valuation headwind")
_sm(ac, "sp500_tr",      "M", "+1", 0.55, "S&P 500 composite price momentum")
_sm(ac, "oas_hy_mom",    "M", "+1", 0.25, "HY spread tightening = risk appetite")
_sm(ac, "cdx_hy_mom",    "M", "+1", 0.20, "CDX HY price up = risk-on")
_sm(ac, "vix",           "S", "+1", 0.30, "Contrarian: high VIX = oversold = buy")
_sm(ac, "aaii_z",        "S", "-1", 0.20, "Contrarian: high bullish = crowded = sell")
_sm(ac, "fci_z",         "S", "-1", 0.15, "Tight FCI = financial headwind")
_sm(ac, "cdx_hy_mom",    "S", "+1", 0.15, "CDX HY risk appetite signal")
_sm(ac, "modern_ted",    "S", "-1", 0.05, "Funding stress = equity headwind")
_sm(ac, "pe_score_sp500","V", "+1", 0.40, "S&P 500 PE valuation score (cheap=+)")
_sm(ac, "erp_us",        "V", "+1", 0.40, "US ERP; high = equities cheap vs bonds")
_sm(ac, "rel_pe_us_em",  "V", "+1", 0.20, "US cheap vs EM = bullish US equity")

# ── US GROWTH ─────────────────────────────────────────────────────────────────
ac = "us_growth"
_sm(ac, "pmi_us",        "F", "+1", 0.35, "Growth = earnings = bullish growth")
_sm(ac, "cesi_us",       "F", "+1", 0.25, "Macro surprise")
_sm(ac, "gdp_us",        "F", "+1", 0.20, "GDP revision")
_sm(ac, "eps_us",        "F", "+1", 0.15, "EPS revisions")
_sm(ac, "breakeven_5y",  "F", "-1", 0.05, "Inflation headwind (rate risk)")
_sm(ac, "sp500_gro_tr",  "M", "+1", 0.70, "S&P 500 Growth TR momentum")
_sm(ac, "cdx_hy_mom",    "M", "+1", 0.30, "Risk appetite = bullish growth")
_sm(ac, "vix",           "S", "+1", 0.30, "Contrarian VIX")
_sm(ac, "aaii_z",        "S", "-1", 0.20, "Contrarian AAII")
_sm(ac, "fci_z",         "S", "-1", 0.15, "Tight FCI headwind")
_sm(ac, "cdx_hy_mom",    "S", "+1", 0.15, "Risk appetite")
_sm(ac, "modern_ted",    "S", "-1", 0.05, "Funding stress")
_sm(ac, "pe_score_gro",  "V", "+1", 0.35, "Growth PE valuation score")
_sm(ac, "erp_us",        "V", "+1", 0.35, "US ERP (growth shares earnings yield)")
_sm(ac, "rel_pe_gro_val","V", "+1", 0.30, "Growth cheap vs value = bullish growth")

# ── US VALUE ──────────────────────────────────────────────────────────────────
ac = "us_value"
_sm(ac, "pmi_us",        "F", "+1", 0.35, "PMI up = cyclical recovery = bullish value")
_sm(ac, "cesi_us",       "F", "+1", 0.25, "Macro surprise")
_sm(ac, "gdp_us",        "F", "+1", 0.20, "GDP growth = cyclical = value positive")
_sm(ac, "eps_us",        "F", "+1", 0.15, "EPS revisions")
_sm(ac, "breakeven_5y",  "F", "-1", 0.05, "Inflation headwind")
_sm(ac, "sp500_val_tr",  "M", "+1", 0.55, "S&P 500 Value TR momentum")
_sm(ac, "oas_bbb_mom",   "M", "+1", 0.25, "Credit tightening = value positive")
_sm(ac, "cdx_ig_mom",    "M", "+1", 0.20, "CDX IG momentum = risk environment")
_sm(ac, "vix",           "S", "+1", 0.30, "Contrarian VIX")
_sm(ac, "aaii_z",        "S", "-1", 0.20, "Contrarian AAII")
_sm(ac, "fci_z",         "S", "-1", 0.15, "Tight FCI headwind")
_sm(ac, "cdx_hy_mom",    "S", "+1", 0.15, "Risk appetite")
_sm(ac, "modern_ted",    "S", "-1", 0.05, "Funding stress")
_sm(ac, "pe_score_val",  "V", "+1", 0.35, "Value PE valuation score")
_sm(ac, "erp_us",        "V", "+1", 0.35, "US ERP")
_sm(ac, "rel_pe_val_gro","V", "+1", 0.30, "Value cheap vs growth = bullish value")

# ── DM EX-US EQUITY ───────────────────────────────────────────────────────────
ac = "dm_equity"
_sm(ac, "pmi_ez",        "F", "+1", 0.25, "EZ PMI composite = DM growth")
_sm(ac, "cesi_ez",       "F", "+1", 0.25, "EZ surprise = DM growth proxy")
_sm(ac, "gdp_dm",        "F", "+1", 0.15, "DM GDP revision")
_sm(ac, "eps_eafe",      "F", "+1", 0.15, "EAFE EPS revisions")
_sm(ac, "pmi_japan_mfg", "F", "+1", 0.10, "Japan Mfg PMI")
_sm(ac, "gdp_japan",     "F", "+1", 0.05, "Japan GDP blend")
_sm(ac, "eps_japan",     "F", "+1", 0.05, "Japan EPS revisions")
_sm(ac, "eafe_tr",       "M", "+1", 0.65, "EAFE TR composite momentum")
_sm(ac, "msci_acwi_tr",  "M", "+1", 0.35, "ACWI momentum = global trend")
_sm(ac, "vix",           "S", "+1", 0.30, "Contrarian VIX = global risk-off reversal")
_sm(ac, "vstoxx_z",      "S", "+1", 0.25, "EZ vol contrarian")
_sm(ac, "cdx_hy_mom",    "S", "+1", 0.30, "Risk appetite = DM equity tailwind")
_sm(ac, "modern_ted",    "S", "-1", 0.15, "Funding stress = EM DM headwind")
_sm(ac, "pe_score_eafe", "V", "+1", 0.35, "EAFE PE valuation score")
_sm(ac, "erp_acwi",      "V", "+1", 0.35, "ACWI ERP; high = DM equities cheap vs bonds")
_sm(ac, "rel_pe_dm_us",  "V", "+1", 0.30, "DM cheap vs US = bullish DM equity")

# ── EM EQUITY ─────────────────────────────────────────────────────────────────
ac = "em_equity"
_sm(ac, "pmi_china",     "F", "+1", 0.30, "China PMI = EM growth engine")
_sm(ac, "cesi_em",       "F", "+1", 0.25, "EM macro surprises")
_sm(ac, "gdp_em",        "F", "+1", 0.25, "EM GDP consensus")
_sm(ac, "eps_em",        "F", "+1", 0.20, "EM EPS revisions")
_sm(ac, "msci_em_tr",    "M", "+1", 0.45, "MSCI EM price momentum")
_sm(ac, "em_xchina_tr",  "M", "+1", 0.30, "EM ex-China breadth check")
_sm(ac, "oas_em_mom",    "M", "+1", 0.25, "EM spread tightening = risk-on")
_sm(ac, "dxy_z",         "S", "-1", 0.30, "Strong USD = EM headwind")
_sm(ac, "embi",          "S", "-1", 0.25, "EM sovereign stress")
_sm(ac, "vix",           "S", "+1", 0.25, "Contrarian VIX for EM equity")
_sm(ac, "em_stress",     "S", "-1", 0.20, "EM credit stress = bearish EM")
_sm(ac, "pe_score_em",   "V", "+1", 0.30, "MSCI EM PE valuation score")
_sm(ac, "erp_em",        "V", "+1", 0.30, "EM ERP; high = cheap vs bonds")
_sm(ac, "rel_pe_em_us",  "V", "+1", 0.25, "EM cheap vs US = bullish EM")
_sm(ac, "oas_em",        "V", "+1", 0.15, "EM Corp OAS as valuation context")

# ── EM EX-CHINA ───────────────────────────────────────────────────────────────
ac = "em_xchina"
# F, M, S: same as EM Equity
_sm(ac, "pmi_china",     "F", "+1", 0.30, "China PMI = EM ex-China growth proxy")
_sm(ac, "cesi_em",       "F", "+1", 0.25, "EM macro surprises")
_sm(ac, "gdp_em",        "F", "+1", 0.25, "EM GDP consensus")
_sm(ac, "eps_em",        "F", "+1", 0.20, "EM EPS revisions")
_sm(ac, "msci_em_tr",    "M", "+1", 0.45, "MSCI EM momentum")
_sm(ac, "em_xchina_tr",  "M", "+1", 0.30, "EM ex-China direct momentum")
_sm(ac, "oas_em_mom",    "M", "+1", 0.25, "EM spread momentum")
_sm(ac, "dxy_z",         "S", "-1", 0.30, "Strong USD = EM headwind")
_sm(ac, "embi",          "S", "-1", 0.25, "EM sovereign stress")
_sm(ac, "vix",           "S", "+1", 0.25, "Contrarian VIX")
_sm(ac, "em_stress",     "S", "-1", 0.20, "EM credit stress")
_sm(ac, "pe_score_emx",  "V", "+1", 0.35, "EM ex-China PE valuation score")
_sm(ac, "erp_em_xchina", "V", "+1", 0.35, "EM ex-China ERP")
_sm(ac, "oas_em",        "V", "+1", 0.30, "EM OAS pctile as valuation context")

# ── CHINA EQUITY ──────────────────────────────────────────────────────────────
ac = "china_equity"
_sm(ac, "pmi_china",     "F", "+1", 0.40, "China Caixin PMI composite")
_sm(ac, "cesi_china",    "F", "+1", 0.25, "China macro surprise")
_sm(ac, "gdp_china",     "F", "+1", 0.20, "China GDP blend")
_sm(ac, "eps_china",     "F", "+1", 0.15, "China EPS revisions")
_sm(ac, "china_tr",      "M", "+1", 0.70, "MSCI China TR momentum")
_sm(ac, "oas_em_mom",    "M", "+1", 0.30, "EM spread momentum = EM risk appetite")
_sm(ac, "dxy_z",         "S", "-1", 0.35, "Strong USD = China headwind")
_sm(ac, "em_stress",     "S", "-1", 0.35, "EM credit stress = bad for China equity")
_sm(ac, "vix",           "S", "+1", 0.30, "Contrarian VIX for China equity")
_sm(ac, "pe_score_china","V", "+1", 0.40, "MSCI China PE valuation score")
_sm(ac, "erp_china",     "V", "+1", 0.35, "China ERP; high = cheap vs bonds")
_sm(ac, "rel_pe_china_us","V","+1", 0.25, "China cheap vs US = bullish China")


# ─────────────────────────────────────────────────────────────────────────────
# 3 — TransformCodes SHEET
# ─────────────────────────────────────────────────────────────────────────────
TRANSFORM_CODES = [
    ("ewma_z",    "EWMA Z-Score",                "ewma_zscore(s, span=window); clip ±3σ",                  "FALSE"),
    ("rolling_z", "Rolling Z-Score",             "rolling_zscore(s, window); slow-moving valuation signals", "FALSE"),
    ("pctile",    "Percentile Rank → Z",         "pctile_rank(s,window); rescale (p-0.5)*4 to z-range",    "FALSE"),
    ("mom_z",     "Momentum Z-Score",            "ewma_zscore(s.pct_change(window)); EPS revisions",       "FALSE"),
    ("price_mom", "Composite Price Momentum",    "composite_price_momentum(s); uses MomentumConfig weights","FALSE"),
    ("inv_mom_z", "Inverted Momentum Z-Score",   "-ewma_zscore(s.diff(window)); falling=positive (yields/OAS)","FALSE"),
]

# ─────────────────────────────────────────────────────────────────────────────
# 4 — MomentumConfig SHEET
# ─────────────────────────────────────────────────────────────────────────────
# Weights must sum to 1.0 per row; 0 = disable component.
MOM_CONFIG_HEADERS = [
    "series_id", "asset_type", "notes",
    "ret_1m_weight", "ret_3m_weight", "ret_6m_weight", "ret_12_1m_weight",
    "ma_dist_weight", "rsi_weight", "range_pos_weight",
    "ret_1m_days", "ret_3m_days", "ret_6m_days", "ret_12m_days",
    "skip_days", "ma_fast", "ma_slow", "rsi_period",
]

MOM_CONFIG_ROWS = [
    # Equity: 12-1M dominates (Jegadeesh-Titman academic momentum)
    ("sp500_tr",     "equity", "S&P 500",         0.10, 0.25, 0.00, 0.40, 0.15, 0.10, 0.00, 21, 63, 126, 252, 21, 50, 200, 14),
    ("sp500_gro_tr", "equity", "S&P 500 Growth",  0.10, 0.25, 0.00, 0.40, 0.15, 0.10, 0.00, 21, 63, 126, 252, 21, 50, 200, 14),
    ("sp500_val_tr", "equity", "S&P 500 Value",   0.10, 0.25, 0.00, 0.40, 0.15, 0.10, 0.00, 21, 63, 126, 252, 21, 50, 200, 14),
    ("eafe_tr",      "equity", "MSCI EAFE",        0.10, 0.25, 0.00, 0.40, 0.15, 0.10, 0.00, 21, 63, 126, 252, 21, 50, 200, 14),
    ("msci_em_tr",   "equity", "MSCI EM",          0.10, 0.25, 0.00, 0.40, 0.15, 0.10, 0.00, 21, 63, 126, 252, 21, 50, 200, 14),
    ("em_xchina_tr", "equity", "MSCI EM ex-China", 0.10, 0.25, 0.00, 0.40, 0.15, 0.10, 0.00, 21, 63, 126, 252, 21, 50, 200, 14),
    ("china_tr",     "equity", "MSCI China",       0.10, 0.25, 0.00, 0.40, 0.15, 0.10, 0.00, 21, 63, 126, 252, 21, 50, 200, 14),
    ("msci_acwi_tr", "equity", "MSCI ACWI",        0.10, 0.25, 0.00, 0.40, 0.15, 0.10, 0.00, 21, 63, 126, 252, 21, 50, 200, 14),
    # Fixed Income: short-term returns + MAs more relevant, skip-month less predictive
    ("bfu5_price",   "fi",     "BFU5 1-3Y UST",    0.20, 0.35, 0.00, 0.15, 0.25, 0.05, 0.00, 21, 63, 126, 252, 21, 50, 200, 14),
    ("i132_price",   "fi",     "ICE 1-3Y UST",     0.20, 0.35, 0.00, 0.15, 0.25, 0.05, 0.00, 21, 63, 126, 252, 21, 50, 200, 14),
    ("lt03_price",   "fi",     "LT03 Short Govt",  0.20, 0.35, 0.00, 0.15, 0.25, 0.05, 0.00, 21, 63, 126, 252, 21, 50, 200, 14),
    ("bsgv_price",   "fi",     "BSGV Long Govt",   0.20, 0.35, 0.00, 0.15, 0.25, 0.05, 0.00, 21, 63, 126, 252, 21, 50, 200, 14),
]


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print(f"Opening {XLSX}")
    wb = load_workbook(XLSX)

    # ── Step 1: DataSeries — add columns ──────────────────────────────────────
    ws_ds = wb["DataSeries"]

    # Find current headers in row 1
    headers = [ws_ds.cell(1, c).value for c in range(1, ws_ds.max_column + 1)]
    new_cols = ["series_type", "input_sheet", "input_column", "transform_code"]
    col_offset = ws_ds.max_column + 1

    for i, col_name in enumerate(new_cols):
        if col_name not in headers:
            c = col_offset + i
            cell = ws_ds.cell(1, c, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BORDER
            print(f"  Added column '{col_name}' at position {c}")
        else:
            print(f"  Column '{col_name}' already exists — skipping header add")

    # Refresh headers after possible additions
    headers = [ws_ds.cell(1, c).value for c in range(1, ws_ds.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}

    def _sid_col():
        return col_idx.get("series_id", 1)

    # Get existing series_ids to avoid duplicates
    existing_sids = set()
    for row in ws_ds.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing_sids.add(str(row[0]))

    # Build lookup: series_id -> (series_type, input_sheet, input_column, transform_code)
    exec_lookup = {}
    for row_data in list(EXEC_ORIGINAL) + list(EXEC_CUSTOM):
        sid = row_data[0]
        exec_lookup[sid] = (row_data[9], row_data[10], row_data[11], row_data[12])

    # Find column positions for the new columns in this sheet
    type_col  = col_idx.get("series_type")
    sheet_col = col_idx.get("input_sheet")
    incol_col = col_idx.get("input_column")
    tc_col    = col_idx.get("transform_code")

    # Update existing rows that are missing executable columns
    updated = 0
    for row_num in range(2, ws_ds.max_row + 1):
        sid_cell = ws_ds.cell(row_num, col_idx.get("series_id", 1))
        if not sid_cell.value:
            continue
        sid = str(sid_cell.value)
        if sid not in exec_lookup:
            continue
        # Check if already populated
        if type_col and ws_ds.cell(row_num, type_col).value:
            continue
        stype, isheet, icol, tcode = exec_lookup[sid]
        if type_col:
            ws_ds.cell(row_num, type_col).value  = stype
        if sheet_col:
            ws_ds.cell(row_num, sheet_col).value = isheet
        if incol_col:
            ws_ds.cell(row_num, incol_col).value = icol
        if tc_col:
            ws_ds.cell(row_num, tc_col).value    = tcode
        updated += 1

    print(f"  Updated {updated} existing DataSeries rows with executable columns")

    # Append truly new rows (not yet in DataSeries)
    all_new = list(EXEC_ORIGINAL) + list(EXEC_CUSTOM)
    added = 0
    for row_data in all_new:
        sid = row_data[0]
        if sid in existing_sids:
            continue  # already handled via update above

        new_row = list(row_data)
        ws_ds.append(new_row)
        row_num = ws_ds.max_row
        for c in range(1, len(new_row) + 1):
            ws_ds.cell(row_num, c).fill = GREEN_FILL
        existing_sids.add(sid)
        added += 1

    print(f"  Added {added} new DataSeries rows")
    ws_ds.freeze_panes = "A2"

    # ── Step 2: SignalMapping — append new rows ────────────────────────────────
    ws_sm = wb["SignalMapping"]
    sm_headers = [ws_sm.cell(1, c).value for c in range(1, ws_sm.max_column + 1)]

    # Get existing (ac, series_id, pillar) combos to avoid dups
    existing_sm = set()
    for row in ws_sm.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            existing_sm.add((str(row[0]), str(row[1]), str(row[2])))

    sm_added = 0
    for (ac, sid, pillar, sign, weight, desc) in SM_ROWS:
        key = (ac, sid, pillar)
        if key in existing_sm:
            print(f"  SKIP SignalMapping {ac}/{sid}/{pillar} — exists")
            continue
        ws_sm.append([ac, sid, pillar, sign, weight, desc])
        row_num = ws_sm.max_row
        for c in range(1, 7):
            ws_sm.cell(row_num, c).fill = BLUE_FILL
        existing_sm.add(key)
        sm_added += 1

    print(f"  Added {sm_added} new SignalMapping rows")
    ws_sm.freeze_panes = "A2"

    # ── Step 3: TransformCodes sheet ──────────────────────────────────────────
    if "TransformCodes" not in wb.sheetnames:
        ws_tc = wb.create_sheet("TransformCodes")
        ws_tc.sheet_properties.tabColor = "F59E0B"
        ws_tc.append(["code", "description", "formula_notes", "requires_two_columns"])
        for row in TRANSFORM_CODES:
            ws_tc.append(list(row))
        _hdr(ws_tc, ncols=4)
        _autosize(ws_tc)
        ws_tc.freeze_panes = "A2"
        print("  Created TransformCodes sheet")
    else:
        print("  TransformCodes sheet already exists — skipping")

    # ── Step 4: MomentumConfig sheet ──────────────────────────────────────────
    if "MomentumConfig" not in wb.sheetnames:
        ws_mc = wb.create_sheet("MomentumConfig")
        ws_mc.sheet_properties.tabColor = "A855F7"
        ws_mc.append(MOM_CONFIG_HEADERS)
        for row in MOM_CONFIG_ROWS:
            ws_mc.append(list(row))
        _hdr(ws_mc, ncols=len(MOM_CONFIG_HEADERS))
        _autosize(ws_mc)
        ws_mc.freeze_panes = "A2"
        print("  Created MomentumConfig sheet")
    else:
        print("  MomentumConfig sheet already exists — skipping")

    wb.save(XLSX)
    print(f"\nSaved: {XLSX}")
    print(f"Summary: +{added} DataSeries rows, +{sm_added} SignalMapping rows, "
          f"+TransformCodes, +MomentumConfig")


if __name__ == "__main__":
    main()
